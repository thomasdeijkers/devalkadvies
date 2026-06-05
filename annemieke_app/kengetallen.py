from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class ImportedReferenceLine:
    line_number: int
    project_name: str | None
    relation_name: str | None
    document_date: datetime | None
    omschrijving_werkzaamheden: str
    hoeveelheid: Decimal | None
    eenheid: str | None
    norm_arbeid: Decimal | None
    uren: Decimal | None
    materiaal: Decimal | None
    materieel: Decimal | None
    onderaannemer: Decimal | None
    totaal_prijs_per_regel: Decimal | None
    eenheidsprijs: Decimal | None
    confidence: int
    raw_text: str


HEADER_ALIASES = {
    "project_name": {"project", "projectnaam", "projectomschrijving"},
    "relation_name": {"relatie", "opdrachtgever", "klant", "client"},
    "document_date": {"datum", "documentdatum", "peildatum", "prijsdatum"},
    "omschrijving_werkzaamheden": {"omschrijving", "werkzaamheden", "omschrijving/ werkzaamheden", "post", "activiteit"},
    "hoeveelheid": {"hvh", "hoeveelheid", "aantal"},
    "eenheid": {"ehd", "eenheid", "unit"},
    "norm_arbeid": {"norm", "norm/ arbeid", "norm arbeid", "arbeid"},
    "uren": {"uren", "uur"},
    "materiaal": {"materiaal"},
    "materieel": {"materieel"},
    "onderaannemer": {"o.a.", "oa", "onderaannemer"},
    "totaal_prijs_per_regel": {"totaal", "totaal prijs per regel", "totaalprijs", "eindprijs"},
    "eenheidsprijs": {"eenheidsprijs", "ehprijs", "prijs/eenheid", "prijs per eenheid"},
}


def import_reference_lines(path: Path) -> list[ImportedReferenceLine]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    header_row, mapping = _find_header_mapping(sheet.iter_rows(values_only=True))
    if not mapping:
        return []

    lines: list[ImportedReferenceLine] = []
    for excel_row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if excel_row_number <= header_row:
            continue
        values = {field: _cell(row, index) for field, index in mapping.items()}
        description = _text(values.get("omschrijving_werkzaamheden"))
        if not description:
            continue

        quantity = _decimal(values.get("hoeveelheid"))
        total = _decimal(values.get("totaal_prijs_per_regel"))
        unit_price = _decimal(values.get("eenheidsprijs"))
        if unit_price is None and total is not None and quantity not in {None, Decimal("0")}:
            try:
                unit_price = total / quantity
            except (InvalidOperation, ZeroDivisionError):
                unit_price = None

        confidence = 100 if unit_price is not None else 70 if total is not None else 45
        lines.append(
            ImportedReferenceLine(
                line_number=len(lines) + 1,
                project_name=_text(values.get("project_name")) or None,
                relation_name=_text(values.get("relation_name")) or None,
                document_date=_date(values.get("document_date")),
                omschrijving_werkzaamheden=description,
                hoeveelheid=quantity,
                eenheid=_text(values.get("eenheid")) or None,
                norm_arbeid=_decimal(values.get("norm_arbeid")),
                uren=_decimal(values.get("uren")),
                materiaal=_decimal(values.get("materiaal")),
                materieel=_decimal(values.get("materieel")),
                onderaannemer=_decimal(values.get("onderaannemer")),
                totaal_prijs_per_regel=total,
                eenheidsprijs=unit_price,
                confidence=confidence,
                raw_text=" | ".join(_text(cell) for cell in row if _text(cell)),
            )
        )
    return lines


def _find_header_mapping(rows: Any) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(rows, start=1):
        normalized = [_normalize_header(cell) for cell in row]
        mapping: dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            for index, header in enumerate(normalized):
                if header in aliases:
                    mapping[field] = index
                    break
        if "omschrijving_werkzaamheden" in mapping and ("eenheidsprijs" in mapping or "totaal_prijs_per_regel" in mapping):
            return row_number, mapping
    return 0, {}


def _cell(row: tuple[Any, ...], index: int) -> Any:
    if index >= len(row):
        return None
    return row[index]


def _normalize_header(value: Any) -> str:
    return _text(value).lower().replace("\n", " ").strip()


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    cleaned = str(value).strip().replace("€", "").replace(" ", "")
    if not cleaned:
        return None
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = _text(value)
    for pattern in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None
