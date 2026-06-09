import os
import re
import time
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from html import escape
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from fastapi import BackgroundTasks, Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import and_, func, or_, select, text
from sqlalchemy.orm import Session, aliased

from .config import settings
from .database import SessionLocal, create_db, engine, get_session
from .exporter import budget_document_to_xlsx, control_model_document_to_xlsx, selected_budget_lines_to_xlsx
from .index_provider import sync_price_index_series
from .kengetallen import import_reference_lines
from .models import (
    AssessmentTemplate,
    BudgetLine,
    ExtractedField,
    IncomingDocument,
    NormalizationTerm,
    OpenAIUsageEvent,
    PriceIndexSeries,
    PriceIndexValue,
    Project,
    ReferenceDataset,
    ReferenceLine,
    Relation,
    ScheduledJob,
)
from .normalizer import (
    apply_normalization,
    is_noise_line,
    normalization_key,
    seed_default_normalization_terms,
    split_aliases,
)
from .parser import parse_pdf
from .template_exporter import fill_screening_template


app = FastAPI(title=settings.app_name)
templates = Jinja2Templates(directory=str(Path(__file__).parent / "templates"))
app.mount("/static", StaticFiles(directory=str(Path(__file__).parent / "static")), name="static")
logo_path = Path(__file__).resolve().parent.parent / "logo.webp"
m3e_logo_path = Path(__file__).resolve().parent.parent / "logo_M3E.jpg"
APP_STARTED_AT = time.time()
MONEY_TEXT_PATTERN = re.compile(
    r"(?:€\s*)?-?(?:\d{1,3}(?:[.\s]\d{3})+(?:,\d{2})?|\d{1,3}(?:,\d{3})+(?:\.\d{2})?|\d+(?:[.,]\d{2}))"
)
REFERENCE_INDEX_TYPE = "kengetal_index"
REFERENCE_CATEGORY_COLUMNS: list[tuple[str, str]] = [
    ("algemeen", "Algemeen"),
    ("sloopwerk", "Sloopwerk"),
    ("fundering", "Fundering"),
    ("skelet", "Skelet"),
    ("daken", "Daken"),
    ("gevel", "Gevel"),
    ("binnenwanden", "Binnenwanden"),
    ("vloerafwerking", "Vloerafwerking"),
    ("trappen", "Trappen"),
    ("plafondafwerking", "Plafondafwerking"),
    ("vaste inrichting", "Vaste inrichting"),
    ("terrein", "Terrein"),
    ("w installatie", "W installatie"),
    ("e installatie", "E installatie"),
    ("t installatie", "T installatie"),
    ("bouwkundig", "Bouwkundig"),
    ("installatie", "Installatie"),
]


def euro(value: Decimal | int | float | str | None) -> str:
    if value is None or value == "":
        return ""
    try:
        amount = Decimal(str(value))
    except InvalidOperation:
        return str(value)
    formatted = f"{amount:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"€ {formatted}"


def amount(value: Decimal | int | float | str | None) -> str:
    if value is None or value == "":
        return ""
    try:
        number = Decimal(str(value))
    except InvalidOperation:
        return str(value)
    if number == number.to_integral_value():
        return str(number.quantize(Decimal("1")))
    formatted = f"{number.quantize(Decimal('0.01')):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return formatted.rstrip("0").rstrip(",")


def quantity(value: Decimal | int | float | str | None) -> str:
    if value is None or value == "":
        return ""
    try:
        number = Decimal(str(value))
    except InvalidOperation:
        return str(value)
    if number == number.to_integral_value():
        return str(number.quantize(Decimal("1")))
    formatted = f"{number.quantize(Decimal('0.1')):,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return formatted.rstrip("0").rstrip(",")


def calculated_unit_price(line: BudgetLine) -> Decimal | None:
    if line.totaal_prijs_per_regel is not None and line.hoeveelheid not in {None, 0}:
        try:
            return line.totaal_prijs_per_regel / line.hoeveelheid
        except (InvalidOperation, ZeroDivisionError):
            return line.eenheidsprijs
    return line.eenheidsprijs or _price_component_total(line)


def calculated_line_total(line: BudgetLine) -> Decimal | None:
    if line.totaal_prijs_per_regel is not None:
        return Decimal(str(line.totaal_prijs_per_regel))
    unit_price = calculated_unit_price(line)
    if unit_price is None:
        return None
    if line.hoeveelheid not in {None, 0}:
        try:
            return Decimal(str(line.hoeveelheid)) * Decimal(str(unit_price))
        except InvalidOperation:
            return None
    return Decimal(str(unit_price))


templates.env.filters["euro"] = euro
templates.env.filters["amount"] = amount
templates.env.filters["quantity"] = quantity
templates.env.filters["unit_price"] = calculated_unit_price
templates.env.filters["line_total"] = calculated_line_total


@app.on_event("startup")
def startup() -> None:
    create_db()
    with SessionLocal() as session:
        if seed_default_normalization_terms(session):
            _reapply_all_normalization(session)


@app.get("/", response_class=HTMLResponse)
def dashboard(
    request: Request,
    project: str | None = None,
    status: str | None = None,
    session: Session = Depends(get_session),
) -> HTMLResponse:
    return _render_workspace(request, session, "overview", project, status)


@app.get("/documents", response_class=HTMLResponse)
def documents_page(
    request: Request,
    project: str | None = None,
    status: str | None = None,
    session: Session = Depends(get_session),
) -> HTMLResponse:
    return _render_workspace(request, session, "documents", project, status)


@app.get("/kengetallen", response_class=HTMLResponse)
def reference_data_page(
    request: Request,
    q: str | None = None,
    session: Session = Depends(get_session),
) -> HTMLResponse:
    return _render_workspace(request, session, "kengetallen", query=q)


@app.get("/beoordeling", response_class=HTMLResponse)
def assessment_page(request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    return _render_workspace(request, session, "beoordeling")


@app.get("/templates", response_class=HTMLResponse)
def templates_page(request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    return _render_workspace(request, session, "templates")


@app.get("/consult", response_class=HTMLResponse)
def consult_page(
    request: Request,
    q: str | None = None,
    status: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    priced: str | None = None,
    min_score: str | None = None,
    session: Session = Depends(get_session),
) -> HTMLResponse:
    return _render_workspace(
        request,
        session,
        "consult",
        status=status,
        query=q,
        date_from=date_from,
        date_to=date_to,
        priced=priced,
        min_score=min_score,
    )


@app.get("/relations", response_class=HTMLResponse)
def relations_page(request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    return _render_workspace(request, session, "relations")


@app.get("/projects", response_class=HTMLResponse)
def projects_page(request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    return _render_workspace(request, session, "projects")


@app.get("/server", response_class=HTMLResponse)
def server_page(request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    return _render_workspace(request, session, "server")


@app.get("/indices", response_class=HTMLResponse)
def indices_page(request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    return _render_workspace(request, session, "indices")


@app.get("/jobs", response_class=HTMLResponse)
def jobs_page(request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    return _render_workspace(request, session, "jobs")


@app.get("/normalisatie", response_class=HTMLResponse)
def normalization_page(request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    return _render_workspace(request, session, "normalisatie")


def _render_workspace(
    request: Request,
    session: Session,
    active_page: str,
    project: str | None = None,
    status: str | None = None,
    query: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    priced: str | None = None,
    min_score: str | None = None,
) -> HTMLResponse:
    document_query = select(IncomingDocument)
    if project:
        search = f"%{project.strip()}%"
        document_query = document_query.outerjoin(IncomingDocument.project).where(
            (IncomingDocument.project_name.ilike(search)) | (Project.name.ilike(search)) | (Project.project_number.ilike(search))
        )
    if status:
        document_query = document_query.where(IncomingDocument.status == status)

    total = session.scalar(select(func.count(IncomingDocument.id))) or 0
    needs_review = session.scalar(
        select(func.count(IncomingDocument.id)).where(IncomingDocument.status == "needs_review")
    ) or 0
    processed = session.scalar(select(func.count(IncomingDocument.id)).where(IncomingDocument.status == "processed")) or 0
    documents = session.scalars(document_query.order_by(IncomingDocument.created_at.desc()).limit(30)).all()
    projects = session.scalars(
        select(Project).order_by(Project.created_at.desc(), Project.name).limit(18)
    ).all()
    relations = session.scalars(
        select(Relation).order_by(Relation.created_at.desc(), Relation.name).limit(18)
    ).all()
    project_options = session.scalars(
        select(Project).order_by(Project.name)
    ).all()
    relation_options = session.scalars(
        select(Relation).order_by(Relation.name)
    ).all()
    index_series = session.scalars(select(PriceIndexSeries).order_by(PriceIndexSeries.name)).all()
    index_values = (
        session.scalars(select(PriceIndexValue).order_by(PriceIndexValue.effective_date.desc()).limit(80)).all()
        if active_page == "indices"
        else []
    )
    latest_index_value = index_values[0] if index_values else None
    scheduled_jobs = session.scalars(select(ScheduledJob).order_by(ScheduledJob.created_at.desc())).all()
    reference_datasets = session.scalars(
        select(ReferenceDataset).order_by(ReferenceDataset.created_at.desc()).limit(20)
    ).all()
    assessment_templates = session.scalars(
        select(AssessmentTemplate).order_by(AssessmentTemplate.created_at.desc()).limit(20)
    ).all()
    reference_line_count = session.scalar(select(func.count(ReferenceLine.id))) or 0
    reference_lines = _reference_lines(session, query) if active_page == "kengetallen" else []
    reference_index_sections = _reference_index_sections(session, query) if active_page == "kengetallen" else []
    reference_index_summary = _reference_index_summary(session, reference_index_sections) if active_page == "kengetallen" else {}
    assessment_documents = session.scalars(
        select(IncomingDocument)
        .where(IncomingDocument.document_type == "begroting_beoordeling")
        .order_by(IncomingDocument.created_at.desc())
        .limit(30)
    ).all()
    normalization_terms = session.scalars(
        select(NormalizationTerm).order_by(NormalizationTerm.canonical_label, NormalizationTerm.alias).limit(400)
    ).all()
    normalization_candidates = _normalization_candidate_lines(session) if active_page == "normalisatie" else []
    normalization_score_levels = _normalization_candidate_score_levels(normalization_candidates)
    normalization_stats = _normalization_stats(session)
    flash_message = _flash_message(request)
    status_context = _status_context(session)
    consult_lines = _consult_lines(session, query, status, date_from, date_to, priced, min_score) if active_page == "consult" else []

    return templates.TemplateResponse(
        request=request,
        name="dashboard.html",
        context={
            "request": request,
            "app_name": settings.app_name,
            "total": total,
            "needs_review": needs_review,
            "processed": processed,
            "documents": documents,
            "projects": projects,
            "relations": relations,
            "project_options": project_options,
            "relation_options": relation_options,
            "index_series": index_series,
            "index_values": index_values,
            "latest_index_value": latest_index_value,
            "scheduled_jobs": scheduled_jobs,
            "reference_datasets": reference_datasets,
            "reference_lines": reference_lines,
            "reference_index_sections": reference_index_sections,
            "reference_index_summary": reference_index_summary,
            "reference_category_columns": REFERENCE_CATEGORY_COLUMNS,
            "assessment_templates": assessment_templates,
            "assessment_documents": assessment_documents,
            "reference_line_count": reference_line_count,
            "normalization_terms": normalization_terms,
            "normalization_candidates": normalization_candidates,
            "normalization_score_levels": normalization_score_levels,
            "normalization_stats": normalization_stats,
            "flash_message": flash_message,
            "selected_project": project or "",
            "selected_query": query or "",
            "selected_status": status or "",
            "selected_date_from": date_from or "",
            "selected_date_to": date_to or "",
            "selected_priced": priced or "",
            "selected_min_score": min_score or "",
            "consult_lines": consult_lines,
            "active_page": active_page,
            **status_context,
        },
    )


@app.get("/logo.webp")
def logo() -> FileResponse:
    return FileResponse(logo_path, media_type="image/webp")


@app.get("/logo-m3e.jpg")
def m3e_logo() -> FileResponse:
    if m3e_logo_path.exists():
        return FileResponse(m3e_logo_path, media_type="image/jpeg")
    return FileResponse(logo_path, media_type="image/webp")


def _document_file_path(document_id: int, session: Session) -> tuple[IncomingDocument, Path]:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")
    file_path = settings.upload_dir / document.stored_filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Bronbestand niet gevonden.")
    return document, file_path


def _document_pdf_path(document_id: int, session: Session) -> tuple[IncomingDocument, Path]:
    document, file_path = _document_file_path(document_id, session)
    if file_path.suffix.lower() != ".pdf":
        raise HTTPException(status_code=400, detail="Dit document is geen PDF.")
    return document, file_path


@app.get("/documents/{document_id}/original")
def original_file(document_id: int, session: Session = Depends(get_session)) -> FileResponse:
    document, file_path = _document_file_path(document_id, session)
    suffix = file_path.suffix.lower()
    media_type = {
        ".pdf": "application/pdf",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
    }.get(suffix, "application/octet-stream")
    return FileResponse(
        file_path,
        media_type=media_type,
        filename=document.original_filename,
        content_disposition_type="inline",
    )


@app.get("/documents/{document_id}/original.pdf")
def original_pdf(document_id: int, session: Session = Depends(get_session)) -> FileResponse:
    document, file_path = _document_pdf_path(document_id, session)
    return FileResponse(
        file_path,
        media_type="application/pdf",
        filename=document.original_filename,
        content_disposition_type="inline",
    )


@app.get("/documents/{document_id}/preview", response_class=HTMLResponse)
def document_preview(document_id: int, session: Session = Depends(get_session)) -> HTMLResponse:
    document, file_path = _document_file_path(document_id, session)
    if file_path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _excel_preview_response(document, file_path)
    if file_path.suffix.lower() != ".pdf":
        return _plain_preview_response(document, "Voor dit bestandstype is nog geen preview beschikbaar.")

    try:
        import fitz
    except Exception as exc:  # pragma: no cover - dependency issue is deployment specific
        raise HTTPException(status_code=500, detail="PDF preview kan niet worden gerenderd.") from exc

    pdf = fitz.open(str(file_path))
    try:
        page_count = pdf.page_count
    finally:
        pdf.close()

    safe_name = escape(document.original_filename)
    pages = "\n".join(
        (
            '<section class="preview-page">'
            f"<span>Pagina {page_number}</span>"
            f'<img loading="lazy" src="/documents/{document_id}/preview/page/{page_number}.png" '
            f'alt="Pagina {page_number} van {safe_name}">'
            "</section>"
        )
        for page_number in range(1, page_count + 1)
    )
    if not pages:
        pages = '<p class="empty">Geen pagina\'s gevonden in dit document.</p>'

    return HTMLResponse(
        f"""<!doctype html>
<html lang="nl">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{safe_name}</title>
    <style>
      :root {{
        color-scheme: dark;
        --bg: #07131b;
        --panel: #0d2132;
        --line: #2b5878;
        --text: #f7fbff;
        --muted: #b8c9d7;
        --accent: #4db1c9;
      }}
      * {{ box-sizing: border-box; }}
      body {{
        margin: 0;
        color: var(--text);
        background: var(--bg);
        font-family: Arial, Helvetica, sans-serif;
      }}
      header {{
        position: sticky;
        top: 0;
        z-index: 2;
        padding: 12px 16px;
        border-bottom: 1px solid var(--line);
        background: rgba(7, 19, 27, 0.96);
      }}
      strong {{
        display: block;
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
        font-size: 14px;
      }}
      small {{
        color: var(--muted);
        font-size: 12px;
      }}
      main {{
        display: grid;
        gap: 18px;
        padding: 18px;
      }}
      .preview-page {{
        display: grid;
        gap: 8px;
        justify-items: center;
      }}
      .preview-page span {{
        justify-self: start;
        color: var(--muted);
        font-size: 12px;
        font-weight: 700;
        text-transform: uppercase;
      }}
      img {{
        width: min(100%, 1180px);
        height: auto;
        border: 1px solid var(--line);
        border-radius: 6px;
        background: #fff;
        box-shadow: 0 16px 40px rgba(0, 0, 0, 0.28);
      }}
      .empty {{
        color: var(--muted);
      }}
    </style>
  </head>
  <body>
    <header>
      <strong>{safe_name}</strong>
      <small>Preview in popup</small>
    </header>
    <main>{pages}</main>
  </body>
</html>"""
    )


def _excel_preview_response(document: IncomingDocument, file_path: Path) -> HTMLResponse:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        sheet = next((item for item in workbook.worksheets if item.sheet_state == "visible"), workbook.worksheets[0])
        rows = []
        max_rows = min(sheet.max_row or 0, 120)
        max_cols = min(sheet.max_column or 0, 28)
        for row in sheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
            if not any(value not in {None, ""} for value in row):
                continue
            rows.append(row)
        workbook.close()
    except Exception as exc:
        return _plain_preview_response(document, f"Excel-preview kon niet worden opgebouwd: {escape(str(exc)[:180])}")

    header_cells = ""
    body_rows = ""
    if rows:
        first_row = rows[0]
        header_cells = "".join(f"<th>{escape(_preview_cell(value))}</th>" for value in first_row)
        body_rows = "\n".join(
            "<tr>" + "".join(f"<td>{escape(_preview_cell(value))}</td>" for value in row) + "</tr>"
            for row in rows[1:]
        )
    else:
        body_rows = '<tr><td class="empty">Geen gevulde cellen gevonden in de eerste regels.</td></tr>'

    safe_name = escape(document.original_filename)
    safe_sheet = escape(sheet.title)
    return HTMLResponse(
        f"""<!doctype html>
<html lang="nl">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{safe_name}</title>
    <style>
      :root {{ color-scheme: dark; --bg: #07131b; --panel: #0d2132; --line: #2b5878; --head: #173e62; --text: #f7fbff; --muted: #b8c9d7; }}
      * {{ box-sizing: border-box; }}
      body {{ margin: 0; color: var(--text); background: var(--bg); font-family: Arial, Helvetica, sans-serif; }}
      header {{ position: sticky; top: 0; z-index: 2; display: flex; justify-content: space-between; gap: 16px; padding: 12px 16px; border-bottom: 1px solid var(--line); background: rgba(7, 19, 27, 0.96); }}
      strong {{ display: block; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; font-size: 14px; }}
      small {{ color: var(--muted); font-size: 12px; }}
      main {{ padding: 16px; overflow: auto; }}
      table {{ min-width: 980px; width: 100%; border-collapse: collapse; background: var(--panel); }}
      th, td {{ max-width: 360px; padding: 7px 9px; border: 1px solid rgba(126, 169, 219, 0.22); vertical-align: top; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; font-size: 12px; }}
      th {{ position: sticky; top: 49px; z-index: 1; color: #d9f4ff; background: var(--head); text-align: left; text-transform: uppercase; font-size: 11px; }}
      td.empty {{ color: var(--muted); }}
    </style>
  </head>
  <body>
    <header>
      <div><strong>{safe_name}</strong><small>Excel-preview: {safe_sheet}</small></div>
      <small>Eerste {len(rows)} gevulde regels</small>
    </header>
    <main>
      <table>
        <thead><tr>{header_cells}</tr></thead>
        <tbody>{body_rows}</tbody>
      </table>
    </main>
  </body>
</html>"""
    )


def _plain_preview_response(document: IncomingDocument, message: str) -> HTMLResponse:
    safe_name = escape(document.original_filename)
    return HTMLResponse(
        f"""<!doctype html>
<html lang="nl"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>{safe_name}</title><style>body{{margin:0;padding:24px;background:#07131b;color:#f7fbff;font-family:Arial,sans-serif}}article{{padding:18px;border:1px solid #2b5878;border-radius:8px;background:#0d2132}}small{{color:#b8c9d7}}</style></head>
<body><article><strong>{safe_name}</strong><p>{message}</p><small>Gebruik download/openen alleen buiten deze popup als je het originele bestand nodig hebt.</small></article></body></html>"""
    )


def _preview_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, Decimal):
        return amount(value)
    return str(value)


@app.get("/documents/{document_id}/preview/page/{page_number}.png")
def document_preview_page(
    document_id: int,
    page_number: int,
    session: Session = Depends(get_session),
) -> StreamingResponse:
    _, file_path = _document_pdf_path(document_id, session)
    try:
        import fitz
    except Exception as exc:  # pragma: no cover - dependency issue is deployment specific
        raise HTTPException(status_code=500, detail="PDF preview kan niet worden gerenderd.") from exc

    pdf = fitz.open(str(file_path))
    try:
        if page_number < 1 or page_number > pdf.page_count:
            raise HTTPException(status_code=404, detail="Pagina niet gevonden.")
        page = pdf.load_page(page_number - 1)
        pixmap = page.get_pixmap(matrix=fitz.Matrix(1.7, 1.7), alpha=False)
        stream = BytesIO(pixmap.tobytes("png"))
    finally:
        pdf.close()

    return StreamingResponse(
        stream,
        media_type="image/png",
        headers={"Cache-Control": "private, max-age=600"},
    )


@app.post("/documents", response_class=HTMLResponse)
async def upload_document(
    background_tasks: BackgroundTasks,
    project_name: str = Form(""),
    project_id: str = Form(""),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Upload een PDF-bestand.")

    settings.upload_dir.mkdir(parents=True, exist_ok=True)
    stored_filename = f"{uuid4().hex}.pdf"
    target_path = settings.upload_dir / stored_filename
    target_path.write_bytes(await file.read())

    selected_project = session.get(Project, _int_or_none(project_id)) if _int_or_none(project_id) else None
    document = IncomingDocument(
        original_filename=file.filename,
        stored_filename=stored_filename,
        project_id=selected_project.id if selected_project else None,
        project_name=project_name.strip() or None,
        status="processing",
        source="upload",
        parsed_text="",
        parser_notes="Upload ontvangen. Parser loopt op de achtergrond.",
        parser_stage="Upload ontvangen",
        parser_progress=5,
    )

    session.add(document)
    session.commit()
    background_tasks.add_task(_parse_document_background, document.id, "budget_upload")

    return RedirectResponse(f"/documents/{document.id}", status_code=303)


@app.post("/kengetallen/upload")
async def upload_reference_dataset(
    name: str = Form(...),
    source: str = Form(""),
    notes: str = Form(""),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="Upload een Excelbestand (.xlsx of .xlsm).")

    settings.upload_dir.mkdir(parents=True, exist_ok=True)
    stored_filename = f"{uuid4().hex}{suffix}"
    target_path = settings.upload_dir / stored_filename
    target_path.write_bytes(await file.read())
    imported_lines = import_reference_lines(target_path)

    dataset = ReferenceDataset(
        name=name.strip(),
        source=source.strip() or None,
        notes=notes.strip() or None,
        original_filename=file.filename,
        stored_filename=stored_filename,
        status="active",
    )
    dataset.lines = [
        ReferenceLine(
            line_number=line.line_number,
            regel_type=line.regel_type,
            niveau=line.niveau,
            hoofdstuk_code=line.hoofdstuk_code,
            hoofdstuk_omschrijving=line.hoofdstuk_omschrijving,
            post_code=line.post_code,
            project_name=line.project_name,
            relation_name=line.relation_name,
            document_date=line.document_date,
            omschrijving_werkzaamheden=line.omschrijving_werkzaamheden,
            hoeveelheid=line.hoeveelheid,
            eenheid=line.eenheid,
            norm_arbeid=line.norm_arbeid,
            uren=line.uren,
            materiaal=line.materiaal,
            materieel=line.materieel,
            onderaannemer=line.onderaannemer,
            totaal_prijs_per_regel=line.totaal_prijs_per_regel,
            eenheidsprijs=line.eenheidsprijs,
            bron_pagina=line.bron_pagina,
            confidence=line.confidence,
            raw_text=line.raw_text,
        )
        for line in imported_lines
        if not is_noise_line(line.omschrijving_werkzaamheden)
    ]
    apply_normalization(session, dataset.lines)
    session.add(dataset)
    session.commit()
    return RedirectResponse("/kengetallen", status_code=303)


@app.post("/templates")
async def create_assessment_template(
    name: str = Form(...),
    version: str = Form(""),
    description: str = Form(""),
    required_columns: str = Form(""),
    output_filename: str = Form(""),
    target_sheet: str = Form(""),
    file: UploadFile | None = File(None),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    default_columns = (
        "omschrijving_werkzaamheden,hoeveelheid,eenheid,norm_arbeid,uren,materiaal,materieel,"
        "onderaannemer,eenheidsprijs,totaal_prijs_per_regel"
    )
    stored_filename = None
    original_filename = None
    if file and file.filename:
        suffix = Path(file.filename).suffix.lower()
        if suffix not in {".xlsx", ".xlsm"}:
            raise HTTPException(status_code=400, detail="Upload een Excel template (.xlsx of .xlsm).")
        settings.upload_dir.mkdir(parents=True, exist_ok=True)
        stored_filename = f"{uuid4().hex}{suffix}"
        (settings.upload_dir / stored_filename).write_bytes(await file.read())
        original_filename = file.filename

    session.add(
        AssessmentTemplate(
            name=name.strip(),
            version=version.strip() or None,
            original_filename=original_filename,
            stored_filename=stored_filename,
            target_sheet=target_sheet.strip() or None,
            description=description.strip() or None,
            required_columns=required_columns.strip() or default_columns,
            output_filename=output_filename.strip() or None,
        )
    )
    session.commit()
    return RedirectResponse("/templates", status_code=303)


@app.post("/kengetallen/{dataset_id}/delete")
def delete_reference_dataset(dataset_id: int, session: Session = Depends(get_session)) -> RedirectResponse:
    dataset = session.get(ReferenceDataset, dataset_id)
    if dataset is None:
        raise HTTPException(status_code=404, detail="Kengetallendataset niet gevonden.")

    file_path = settings.upload_dir / dataset.stored_filename if dataset.stored_filename else None
    session.delete(dataset)
    session.commit()
    if file_path and file_path.exists():
        file_path.unlink()
    return RedirectResponse("/kengetallen", status_code=303)


@app.post("/kengetallen/index-project")
async def update_reference_index_project(
    request: Request,
    session: Session = Depends(get_session),
) -> RedirectResponse:
    form = await request.form()
    dataset_id = _int_or_none(str(form.get("dataset_id", "")))
    source_row = _int_or_none(str(form.get("source_row", "")))
    if dataset_id is None or source_row is None:
        raise HTTPException(status_code=400, detail="Kengetalproject mist dataset of bronrij.")

    dataset = session.get(ReferenceDataset, dataset_id)
    if dataset is None:
        raise HTTPException(status_code=404, detail="Kengetallendataset niet gevonden.")

    lines = session.scalars(
        select(ReferenceLine)
        .where(ReferenceLine.dataset_id == dataset_id)
        .where(ReferenceLine.source_row == source_row)
        .where(ReferenceLine.regel_type == REFERENCE_INDEX_TYPE)
    ).all()
    if not lines:
        line_ids = [
            _int_or_none(str(value))
            for value in form.getlist("line_ids")
            if _int_or_none(str(value)) is not None
        ]
        if line_ids:
            lines = session.scalars(
                select(ReferenceLine)
                .where(ReferenceLine.dataset_id == dataset_id)
                .where(ReferenceLine.id.in_(line_ids))
                .where(ReferenceLine.regel_type == REFERENCE_INDEX_TYPE)
            ).all()
    if not lines:
        raise HTTPException(status_code=404, detail="Kengetalproject niet gevonden.")

    section = str(form.get("section", "") or lines[0].hoofdstuk_omschrijving or "Overig").strip()
    project_name = str(form.get("project_name", "") or "").strip()
    variant = str(form.get("variant", "") or "").strip() or None
    phase = str(form.get("phase", "") or "").strip() or None
    period = str(form.get("period", "") or "").strip() or None
    project_sheet_name = str(form.get("project_sheet_name", "") or "").strip() or None
    document_date = _date_or_none(str(form.get("document_date", "") or ""))
    bdb_indexering = _decimal_or_none(str(form.get("bdb_indexering", "") or ""))
    line_by_category = {(line.post_code or line.hoofdstuk_code or "").lower(): line for line in lines}
    max_line_number = session.scalar(select(func.max(ReferenceLine.line_number)).where(ReferenceLine.dataset_id == dataset_id)) or 0
    changed_lines: list[ReferenceLine] = []

    for category_key, category_label in REFERENCE_CATEGORY_COLUMNS:
        safe_category_key = category_key.replace(" ", "_")
        value = _decimal_or_none(
            str(form.get(f"cat_{safe_category_key}", form.get(f"cat_{category_key}", "")) or "")
        )
        line = line_by_category.get(category_key)
        if value is None:
            if line is not None:
                session.delete(line)
            continue
        if line is None:
            max_line_number += 1
            line = ReferenceLine(
                dataset_id=dataset.id,
                line_number=max_line_number,
                regel_type=REFERENCE_INDEX_TYPE,
                niveau=0,
                hoofdstuk_code=category_key,
                post_code=category_key,
                hoeveelheid=Decimal("1"),
                eenheid="m2 bvo",
                confidence=100,
            )
            session.add(line)
        line.hoofdstuk_code = category_key
        line.hoofdstuk_omschrijving = section
        line.post_code = category_key
        line.project_name = project_name or None
        line.relation_name = variant
        line.document_date = document_date
        line.phase = phase
        line.period = period
        line.bdb_indexering = bdb_indexering
        line.project_sheet_name = project_sheet_name
        line.source_row = source_row
        line.omschrijving_werkzaamheden = f"{category_label} - {section}"
        line.hoeveelheid = Decimal("1")
        line.eenheid = "m2 bvo"
        line.norm_arbeid = bdb_indexering
        line.eenheidsprijs = value
        line.totaal_prijs_per_regel = value
        line.confidence = 100
        if not line.raw_text:
            line.raw_text = f"{dataset.original_filename or dataset.name}|rij:{source_row}|categorie:{category_key}"
        changed_lines.append(line)

    if changed_lines:
        apply_normalization(session, changed_lines)
    session.commit()
    return RedirectResponse("/kengetallen?notice=kengetal_project_opgeslagen#kengetal-index", status_code=303)


@app.post("/normalisatie/terms")
def create_normalization_term(
    canonical_label: str = Form(...),
    aliases: str = Form(""),
    category: str = Form("omschrijving"),
    match_type: str = Form("fuzzy"),
    min_score: str = Form("82"),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    _add_normalization_terms(session, canonical_label, aliases, category, match_type, min_score)
    _reapply_all_normalization(session)
    return RedirectResponse("/normalisatie", status_code=303)


@app.post("/normalisatie/terms/{term_id}")
def update_normalization_term(
    term_id: int,
    canonical_label: str = Form(...),
    alias: str = Form(...),
    category: str = Form("omschrijving"),
    match_type: str = Form("fuzzy"),
    min_score: str = Form("82"),
    active: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    term = session.get(NormalizationTerm, term_id)
    if term is None:
        raise HTTPException(status_code=404, detail="Normalisatieterm niet gevonden.")
    label = canonical_label.strip()
    alias_value = alias.strip()
    if not label or not alias_value:
        raise HTTPException(status_code=400, detail="Vul standaardterm en synoniem in.")
    term.canonical_label = label
    term.canonical_key = normalization_key(label) or label.lower().replace(" ", "_")[:180]
    term.alias = alias_value
    term.category = category.strip() or "omschrijving"
    term.match_type = match_type if match_type in {"hard", "fuzzy"} else "fuzzy"
    term.min_score = max(50, min(100, _int_or_none(min_score) or 82))
    term.active = 1 if active else 0
    _reapply_all_normalization(session)
    return RedirectResponse("/normalisatie", status_code=303)


@app.post("/normalisatie/terms/{term_id}/delete")
def delete_normalization_term(term_id: int, session: Session = Depends(get_session)) -> RedirectResponse:
    term = session.get(NormalizationTerm, term_id)
    if term is None:
        raise HTTPException(status_code=404, detail="Normalisatieterm niet gevonden.")
    session.delete(term)
    _reapply_all_normalization(session)
    return RedirectResponse("/normalisatie", status_code=303)


@app.post("/normalisatie/apply")
def apply_normalization_to_existing(session: Session = Depends(get_session)) -> RedirectResponse:
    _reapply_all_normalization(session)
    return RedirectResponse("/normalisatie?notice=normalisatie_bijgewerkt", status_code=303)


@app.post("/normalisatie/suggesties/{line_id}/validate")
def validate_normalization_suggestion(line_id: int, session: Session = Depends(get_session)) -> RedirectResponse:
    line = session.get(BudgetLine, line_id)
    if line is None:
        raise HTTPException(status_code=404, detail="Voorstel niet gevonden.")
    canonical_label = str(
        line.normalization_candidate or line.normalized_omschrijving or line.omschrijving_werkzaamheden or ""
    ).strip()
    if not canonical_label:
        raise HTTPException(status_code=400, detail="Geen voorstel beschikbaar.")
    _add_normalization_terms(
        session,
        canonical_label,
        line.omschrijving_werkzaamheden,
        "omschrijving",
        "hard",
        "100",
    )
    _reapply_all_normalization(session)
    return RedirectResponse("/normalisatie?notice=voorstel_gevalideerd#woordenboek", status_code=303)


@app.post("/normalisatie/suggesties/{line_id}/terms")
def create_term_from_suggestion(
    line_id: int,
    canonical_label: str = Form(...),
    aliases: str = Form(""),
    category: str = Form("omschrijving"),
    match_type: str = Form("fuzzy"),
    min_score: str = Form("82"),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    line = session.get(BudgetLine, line_id)
    if line is None:
        raise HTTPException(status_code=404, detail="Voorstel niet gevonden.")
    alias_values = aliases
    if line.omschrijving_werkzaamheden and line.omschrijving_werkzaamheden not in alias_values:
        alias_values = f"{line.omschrijving_werkzaamheden}\n{alias_values}".strip()
    _add_normalization_terms(session, canonical_label, alias_values, category, match_type, min_score)
    _reapply_all_normalization(session)
    return RedirectResponse("/normalisatie?notice=term_toegevoegd#woordenboek", status_code=303)


@app.post("/beoordeling/upload")
async def upload_assessment_input(
    background_tasks: BackgroundTasks,
    project_name: str = Form(""),
    project_id: str = Form(""),
    template_id: str = Form(""),
    reference_dataset_id: str = Form(""),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in {".pdf", ".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="Upload een PDF of Excelbestand.")

    settings.upload_dir.mkdir(parents=True, exist_ok=True)
    stored_filename = f"{uuid4().hex}{suffix}"
    target_path = settings.upload_dir / stored_filename
    target_path.write_bytes(await file.read())
    selected_project = session.get(Project, _int_or_none(project_id)) if _int_or_none(project_id) else None

    document = IncomingDocument(
        original_filename=file.filename or "begroting",
        stored_filename=stored_filename,
        project_id=selected_project.id if selected_project else None,
        project_name=project_name.strip() or None,
        status="processing" if suffix == ".pdf" else "needs_review",
        document_type="begroting_beoordeling",
        source="excel" if suffix in {".xlsx", ".xlsm"} else "pdf",
        parser_notes=(
            f"{_assessment_note(template_id, reference_dataset_id)} | Upload ontvangen. Parser loopt op de achtergrond."
            if suffix == ".pdf"
            else _assessment_note(template_id, reference_dataset_id)
        ),
        parser_stage="Upload ontvangen" if suffix == ".pdf" else "Excel verwerkt",
        parser_progress=5 if suffix == ".pdf" else 100,
    )
    if suffix == ".pdf":
        session.add(document)
        session.commit()
        background_tasks.add_task(_parse_document_background, document.id, "assessment_upload", True)
        return RedirectResponse(f"/documents/{document.id}", status_code=303)
    else:
        imported_lines = import_reference_lines(target_path)
        document.parsed_text = "\n".join(line.raw_text for line in imported_lines[:500])
        document.budget_lines = [
            BudgetLine(
                line_number=line.line_number,
                regel_type=line.regel_type,
                niveau=line.niveau,
                hoofdstuk_code=line.hoofdstuk_code,
                hoofdstuk_omschrijving=line.hoofdstuk_omschrijving,
                post_code=line.post_code,
                omschrijving_werkzaamheden=line.omschrijving_werkzaamheden,
                hoeveelheid=line.hoeveelheid,
                eenheid=line.eenheid,
                norm_arbeid=line.norm_arbeid,
                uren=line.uren,
                materiaal=line.materiaal,
                materieel=line.materieel,
                onderaannemer=line.onderaannemer,
                totaal_prijs_per_regel=line.totaal_prijs_per_regel,
                eenheidsprijs=line.eenheidsprijs,
                bron_pagina=line.bron_pagina,
                confidence=line.confidence,
                raw_text=line.raw_text,
            )
            for line in imported_lines
            if not is_noise_line(line.omschrijving_werkzaamheden)
        ]

    for line in document.budget_lines:
        _normalize_line_prices(line)
    apply_normalization(session, document.budget_lines)
    source_total, source_label = _source_total_from_document(document, include_saved=False)
    document.source_total_amount = source_total
    document.source_total_source = source_label
    session.add(document)
    session.commit()
    return RedirectResponse(f"/documents/{document.id}", status_code=303)


@app.post("/relations")
def create_relation(
    relation_type: str = Form("opdrachtgever"),
    name: str = Form(...),
    contact_name: str = Form(""),
    address: str = Form(""),
    postal_code: str = Form(""),
    city: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    website: str = Form(""),
    notes: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    session.add(
        Relation(
            relation_type=relation_type.strip() or "opdrachtgever",
            name=name.strip(),
            contact_name=contact_name.strip() or None,
            address=address.strip() or None,
            postal_code=postal_code.strip() or None,
            city=city.strip() or None,
            phone=phone.strip() or None,
            email=email.strip() or None,
            website=website.strip() or None,
            notes=notes.strip() or None,
        )
    )
    session.commit()
    return RedirectResponse("/relations", status_code=303)


@app.post("/relations/{relation_id}")
def update_relation(
    relation_id: int,
    relation_type: str = Form("opdrachtgever"),
    name: str = Form(...),
    contact_name: str = Form(""),
    address: str = Form(""),
    postal_code: str = Form(""),
    city: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    website: str = Form(""),
    notes: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    relation = session.get(Relation, relation_id)
    if relation is None:
        raise HTTPException(status_code=404, detail="Relatie niet gevonden.")
    relation.relation_type = relation_type.strip() or "opdrachtgever"
    relation.name = name.strip()
    relation.contact_name = contact_name.strip() or None
    relation.address = address.strip() or None
    relation.postal_code = postal_code.strip() or None
    relation.city = city.strip() or None
    relation.phone = phone.strip() or None
    relation.email = email.strip() or None
    relation.website = website.strip() or None
    relation.notes = notes.strip() or None
    session.commit()
    return RedirectResponse(f"/relations#relation-{relation.id}", status_code=303)


@app.post("/projects")
def create_project(
    project_number: str = Form(""),
    name: str = Form(...),
    description: str = Form(""),
    location: str = Form(""),
    status: str = Form("actief"),
    client_relation_id: str = Form(""),
    architect_relation_id: str = Form(""),
    constructor_relation_id: str = Form(""),
    notes: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    session.add(
        Project(
            project_number=project_number.strip() or None,
            name=name.strip(),
            description=description.strip() or None,
            location=location.strip() or None,
            status=status.strip() or "actief",
            client_relation_id=_int_or_none(client_relation_id),
            architect_relation_id=_int_or_none(architect_relation_id),
            constructor_relation_id=_int_or_none(constructor_relation_id),
            notes=notes.strip() or None,
        )
    )
    session.commit()
    return RedirectResponse("/projects", status_code=303)


@app.post("/projects/{project_id}")
def update_project(
    project_id: int,
    project_number: str = Form(""),
    name: str = Form(...),
    description: str = Form(""),
    location: str = Form(""),
    status: str = Form("actief"),
    client_relation_id: str = Form(""),
    architect_relation_id: str = Form(""),
    constructor_relation_id: str = Form(""),
    notes: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    project = session.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project niet gevonden.")
    project.project_number = project_number.strip() or None
    project.name = name.strip()
    project.description = description.strip() or None
    project.location = location.strip() or None
    project.status = status.strip() or "actief"
    project.client_relation_id = _int_or_none(client_relation_id)
    project.architect_relation_id = _int_or_none(architect_relation_id)
    project.constructor_relation_id = _int_or_none(constructor_relation_id)
    project.notes = notes.strip() or None
    session.commit()
    return RedirectResponse(f"/projects#project-{project.id}", status_code=303)


@app.post("/indices")
def create_index_series(
    name: str = Form(...),
    description: str = Form(""),
    source: str = Form(""),
    provider: str = Form("cbs"),
    api_url: str = Form(""),
    period_field: str = Form(""),
    value_field: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    series = PriceIndexSeries(
        name=name.strip(),
        description=description.strip() or None,
        source=source.strip() or None,
        provider=provider.strip() or "manual",
        api_url=api_url.strip() or None,
        period_field=period_field.strip() or None,
        value_field=value_field.strip() or None,
    )
    session.add(series)
    session.commit()
    return RedirectResponse("/indices", status_code=303)


@app.post("/indices/value")
def create_index_value(
    period: str = Form(...),
    index_value: str = Form(...),
    note: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    series = session.scalars(select(PriceIndexSeries).order_by(PriceIndexSeries.created_at).limit(1)).first()
    if series is None:
        series = PriceIndexSeries(
            name="BDB index",
            source="Handmatig",
            provider="manual",
            description="Handmatig beheerde indexwaarden voor begrotingsbeoordeling.",
        )
        session.add(series)
        session.flush()
    parsed_value = _decimal_or_none(index_value)
    if parsed_value is None:
        return RedirectResponse("/indices?notice=index_ongeldig", status_code=303)
    effective_date = _period_to_date(period)
    if effective_date is None:
        return RedirectResponse("/indices?notice=index_periode_ongeldig", status_code=303)
    label = period.strip()
    existing = session.scalar(
        select(PriceIndexValue)
        .where(PriceIndexValue.series_id == series.id)
        .where(PriceIndexValue.notes == label)
        .limit(1)
    )
    if existing is None:
        existing = PriceIndexValue(series_id=series.id, effective_date=effective_date, notes=label)
        session.add(existing)
    existing.index_value = parsed_value
    existing.effective_date = effective_date
    existing.source_reference = note.strip() or None
    session.commit()
    return RedirectResponse("/indices?notice=index_opgeslagen", status_code=303)


@app.post("/indices/value/{value_id}")
def update_index_value(
    value_id: int,
    period: str = Form(...),
    index_value: str = Form(...),
    note: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    value = session.get(PriceIndexValue, value_id)
    if value is None:
        raise HTTPException(status_code=404, detail="Indexregel niet gevonden.")
    parsed_value = _decimal_or_none(index_value)
    if parsed_value is None:
        return RedirectResponse("/indices?notice=index_ongeldig", status_code=303)
    effective_date = _period_to_date(period)
    if effective_date is None:
        return RedirectResponse("/indices?notice=index_periode_ongeldig", status_code=303)
    value.notes = period.strip()
    value.index_value = parsed_value
    value.effective_date = effective_date
    value.source_reference = note.strip() or None
    session.commit()
    return RedirectResponse("/indices?notice=index_opgeslagen", status_code=303)


@app.post("/indices/{series_id}/sync")
def sync_index_series(series_id: int, session: Session = Depends(get_session)) -> RedirectResponse:
    series = session.get(PriceIndexSeries, series_id)
    if series is None:
        raise HTTPException(status_code=404, detail="Indexreeks niet gevonden.")
    sync_price_index_series(session, series)
    return RedirectResponse("/indices", status_code=303)


@app.post("/jobs")
def create_scheduled_job(
    name: str = Form(...),
    job_type: str = Form("index_sync"),
    cron_expression: str = Form("0 5 1 * *"),
    target_id: str = Form(""),
    enabled: str = Form("1"),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    session.add(
        ScheduledJob(
            name=name.strip(),
            job_type=job_type.strip() or "index_sync",
            cron_expression=cron_expression.strip() or "0 5 1 * *",
            target_id=_int_or_none(target_id),
            enabled=1 if enabled else 0,
        )
    )
    session.commit()
    return RedirectResponse("/jobs", status_code=303)


@app.post("/jobs/{job_id}/run")
def run_scheduled_job(job_id: int, session: Session = Depends(get_session)) -> RedirectResponse:
    job = session.get(ScheduledJob, job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Cronjob niet gevonden.")
    try:
        count = _run_job(session, job)
        job.last_status = "ok"
        job.last_message = f"{count} waarden bijgewerkt"
    except Exception as exc:
        job.last_status = "fout"
        job.last_message = str(exc)[:500]
    job.last_run_at = datetime.now(timezone.utc)
    session.commit()
    return RedirectResponse("/jobs", status_code=303)


@app.post("/documents/{document_id}/reparse-openai")
def reparse_document_with_openai(
    background_tasks: BackgroundTasks,
    document_id: int,
    session: Session = Depends(get_session),
) -> RedirectResponse:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")

    document.status = "processing"
    document.parser_notes = "OpenAI herparse gestart. Parser loopt op de achtergrond."
    document.parser_stage = "OpenAI herparse gestart"
    document.parser_progress = 5
    session.commit()
    background_tasks.add_task(_parse_document_background, document.id, "budget_reparse", True)
    return RedirectResponse(f"/documents/{document.id}", status_code=303)


@app.get("/documents/{document_id}", response_class=HTMLResponse)
def document_detail(
    document_id: int,
    request: Request,
    session: Session = Depends(get_session),
) -> HTMLResponse:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")

    if _normalize_document_line_prices(document):
        session.commit()
    budget_line_groups = _budget_line_groups(document.budget_lines)
    assessment_templates = session.scalars(
        select(AssessmentTemplate).where(AssessmentTemplate.status == "active").order_by(AssessmentTemplate.created_at.desc())
    ).all()
    return templates.TemplateResponse(
        request=request,
        name="document_detail.html",
        context={
            "request": request,
            "app_name": settings.app_name,
            "document": document,
            "budget_line_groups": budget_line_groups,
            "total_context": _document_total_context(document),
            "assessment_templates": assessment_templates,
            "selected_template": _template_for_document(session, document),
            "flash_message": _flash_message(request),
            "project_options": session.scalars(select(Project).order_by(Project.name)).all(),
            "relation_options": session.scalars(select(Relation).order_by(Relation.name)).all(),
            "normalization_suggestions": [
                line for line in document.budget_lines if line.normalization_candidate and line.normalization_score < 100
            ][:12],
            **_status_context(session),
        },
    )


@app.post("/documents/{document_id}/meta")
def update_document_meta(
    document_id: int,
    project_id: str = Form(""),
    relation_id: str = Form(""),
    project_name: str = Form(""),
    document_type: str = Form(""),
    source: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")

    selected_project = session.get(Project, _int_or_none(project_id)) if _int_or_none(project_id) else None
    selected_relation = session.get(Relation, _int_or_none(relation_id)) if _int_or_none(relation_id) else None
    document.project_id = selected_project.id if selected_project else None
    document.project_name = project_name.strip() or (selected_project.name if selected_project else None)
    document.document_type = document_type.strip() or None
    document.source = source.strip() or None
    if selected_project and selected_relation:
        selected_project.client_relation_id = selected_relation.id
    session.commit()
    return RedirectResponse(f"/documents/{document.id}", status_code=303)


@app.post("/documents/{document_id}/source-total")
def update_document_source_total(
    document_id: int,
    source_total_amount: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")

    parsed_total = _decimal_or_none(source_total_amount)
    document.source_total_amount = parsed_total
    document.source_total_source = "handmatig" if parsed_total is not None else None
    document.source_total_manual = 1 if parsed_total is not None else 0
    session.commit()
    return RedirectResponse(f"/documents/{document.id}?notice=origineel_totaal_opgeslagen", status_code=303)


@app.post("/documents/{document_id}/fields")
def add_field(
    document_id: int,
    field_name: str = Form(...),
    field_value: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")

    session.add(
        ExtractedField(
            document_id=document.id,
            field_name=field_name.strip().lower().replace(" ", "_"),
            field_value=field_value.strip(),
            confidence=100,
            source="manual",
        )
    )
    document.status = "needs_review"
    session.commit()
    return RedirectResponse(f"/documents/{document.id}", status_code=303)


@app.post("/documents/{document_id}/status")
def update_status(
    document_id: int,
    status: str = Form(...),
    return_to: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")
    if status not in {"needs_review", "processed", "archived"}:
        raise HTTPException(status_code=400, detail="Onbekende status.")

    document.status = status
    session.commit()
    target = _safe_return(return_to) or request_url_for_document(document.id, archived=status == "archived")
    return RedirectResponse(target, status_code=303)


@app.post("/documents/{document_id}/delete")
def delete_document(
    document_id: int,
    return_to: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")

    file_path = settings.upload_dir / document.stored_filename
    session.delete(document)
    session.commit()
    if file_path.exists():
        file_path.unlink()
    return RedirectResponse(_safe_return(return_to) or "/", status_code=303)


@app.post("/documents/{document_id}/lines/{line_id}")
def update_budget_line(
    document_id: int,
    line_id: int,
    omschrijving_werkzaamheden: str = Form(""),
    hoeveelheid: str = Form(""),
    eenheid: str = Form(""),
    norm_arbeid: str = Form(""),
    uren: str = Form(""),
    materiaal: str = Form(""),
    materieel: str = Form(""),
    onderaannemer: str = Form(""),
    eenheidsprijs: str = Form(""),
    totaal_prijs_per_regel: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    line = session.get(BudgetLine, line_id)
    if line is None or line.document_id != document_id:
        raise HTTPException(status_code=404, detail="Begrotingsregel niet gevonden.")

    line.omschrijving_werkzaamheden = omschrijving_werkzaamheden.strip()
    line.hoeveelheid = _decimal_or_none(hoeveelheid)
    line.eenheid = eenheid.strip() or None
    line.norm_arbeid = _decimal_or_none(norm_arbeid)
    line.uren = _decimal_or_none(uren)
    line.materiaal = _decimal_or_none(materiaal)
    line.materieel = _decimal_or_none(materieel)
    line.onderaannemer = _decimal_or_none(onderaannemer)
    line.eenheidsprijs = _decimal_or_none(eenheidsprijs)
    line.totaal_prijs_per_regel = _decimal_or_none(totaal_prijs_per_regel)
    _normalize_line_prices(line)
    line.confidence = 100
    apply_normalization(session, [line])
    session.commit()
    return RedirectResponse(f"/documents/{document_id}", status_code=303)


@app.post("/documents/{document_id}/budget-lines/bulk")
async def update_budget_lines_bulk(
    document_id: int,
    request: Request,
    session: Session = Depends(get_session),
) -> RedirectResponse:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")

    form = await request.form()
    line_ids = form.getlist("line_id")
    delete_line_ids = {str(value) for value in form.getlist("delete_line_id")}
    action = str(form.get("action") or "save")
    updated_lines: list[BudgetLine] = []
    for index, raw_line_id in enumerate(line_ids):
        try:
            line_id = int(str(raw_line_id))
        except ValueError:
            continue

        line = session.get(BudgetLine, line_id)
        if line is None or line.document_id != document_id:
            continue
        if action == "delete_selected" and str(line_id) in delete_line_ids:
            session.delete(line)
            continue

        line.omschrijving_werkzaamheden = _form_value(form, "omschrijving_werkzaamheden", index).strip()
        line.hoeveelheid = _decimal_or_none(_form_value(form, "hoeveelheid", index))
        line.eenheid = _form_value(form, "eenheid", index).strip() or None
        line.norm_arbeid = _decimal_or_none(_form_value(form, "norm_arbeid", index))
        line.uren = _decimal_or_none(_form_value(form, "uren", index))
        line.materiaal = _decimal_or_none(_form_value(form, "materiaal", index))
        line.materieel = _decimal_or_none(_form_value(form, "materieel", index))
        line.onderaannemer = _decimal_or_none(_form_value(form, "onderaannemer", index))
        line.eenheidsprijs = _decimal_or_none(_form_value(form, "eenheidsprijs", index))
        line.totaal_prijs_per_regel = _decimal_or_none(_form_value(form, "totaal_prijs_per_regel", index))
        _normalize_line_prices(line)
        line.confidence = 100
        updated_lines.append(line)

    if updated_lines:
        apply_normalization(session, updated_lines)
    if action == "validate":
        document.status = "processed"
        notice = "regels_gevalideerd"
    elif action == "add_to_references":
        added_count = _add_budget_lines_to_references(
            session,
            document,
            [line for line in updated_lines if str(line.id) in delete_line_ids],
        )
        document.status = "needs_review"
        notice = "kengetallen_toegevoegd" if added_count else "geen_kengetallen_geselecteerd"
    elif action != "delete_selected":
        document.status = "needs_review"
        notice = "regels_opgeslagen"
    else:
        notice = "regels_verwijderd"
    session.commit()
    return RedirectResponse(f"/documents/{document_id}?notice={notice}", status_code=303)


@app.post("/documents/{document_id}/lines")
def add_budget_line(
    document_id: int,
    omschrijving_werkzaamheden: str = Form(""),
    hoeveelheid: str = Form(""),
    eenheid: str = Form(""),
    norm_arbeid: str = Form(""),
    uren: str = Form(""),
    materiaal: str = Form(""),
    materieel: str = Form(""),
    onderaannemer: str = Form(""),
    eenheidsprijs: str = Form(""),
    totaal_prijs_per_regel: str = Form(""),
    session: Session = Depends(get_session),
) -> RedirectResponse:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")

    next_line_number = len(document.budget_lines) + 1
    session.add(
        line := BudgetLine(
            document_id=document.id,
            line_number=next_line_number,
            omschrijving_werkzaamheden=omschrijving_werkzaamheden.strip(),
            hoeveelheid=_decimal_or_none(hoeveelheid),
            eenheid=eenheid.strip() or None,
            norm_arbeid=_decimal_or_none(norm_arbeid),
            uren=_decimal_or_none(uren),
            materiaal=_decimal_or_none(materiaal),
            materieel=_decimal_or_none(materieel),
            onderaannemer=_decimal_or_none(onderaannemer),
            eenheidsprijs=_decimal_or_none(eenheidsprijs),
            totaal_prijs_per_regel=_decimal_or_none(totaal_prijs_per_regel),
            confidence=100,
            raw_text="handmatig toegevoegd",
        )
    )
    _normalize_line_prices(line)
    apply_normalization(session, [line])
    session.commit()
    return RedirectResponse(f"/documents/{document.id}", status_code=303)


@app.post("/documents/{document_id}/lines/{line_id}/delete")
def delete_budget_line(
    document_id: int,
    line_id: int,
    session: Session = Depends(get_session),
) -> RedirectResponse:
    line = session.get(BudgetLine, line_id)
    if line is None or line.document_id != document_id:
        raise HTTPException(status_code=404, detail="Begrotingsregel niet gevonden.")

    session.delete(line)
    session.commit()
    return RedirectResponse(f"/documents/{document_id}", status_code=303)


@app.get("/documents/{document_id}/export.xlsx")
def export_document(document_id: int, session: Session = Depends(get_session)) -> StreamingResponse:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")

    if _normalize_document_line_prices(document):
        session.commit()
    stream = budget_document_to_xlsx(document)
    filename = f"begroting-{document.id}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/documents/{document_id}/controlemodel.xlsx")
def export_control_model(document_id: int, session: Session = Depends(get_session)) -> StreamingResponse:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")

    if _normalize_document_line_prices(document):
        session.commit()
    stream = control_model_document_to_xlsx(document)
    filename = f"controlemodel-{document.id}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/documents/{document_id}/screening.xlsx")
def export_screening_template(
    document_id: int,
    template_id: str = "",
    session: Session = Depends(get_session),
) -> StreamingResponse:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")

    if _normalize_document_line_prices(document):
        session.commit()
    template = _template_for_document(session, document, template_id)
    if template and template.stored_filename:
        template_path = settings.upload_dir / template.stored_filename
        if template_path.exists():
            try:
                stream = fill_screening_template(document, template_path, template.target_sheet, m3e_logo_path)
                filename = _excel_filename(template.output_filename or f"screening-{document.id}.xlsx", template_path.suffix)
                media_type = (
                    "application/vnd.ms-excel.sheet.macroEnabled.12"
                    if filename.lower().endswith(".xlsm")
                    else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                return StreamingResponse(
                    stream,
                    media_type=media_type,
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'},
                )
            except Exception:
                pass

    stream = control_model_document_to_xlsx(document)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="screening-fallback-{document.id}.xlsx"'},
    )


@app.get("/consult/export.xlsx")
def export_consult_selection(
    q: str | None = None,
    status: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    priced: str | None = None,
    min_score: str | None = None,
    session: Session = Depends(get_session),
) -> StreamingResponse:
    lines = _consult_lines(session, q, status, date_from, date_to, priced, min_score, limit=5000)
    stream = selected_budget_lines_to_xlsx(lines)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="raadplegen-selectie.xlsx"'},
    )


@app.get("/api/documents/{document_id}")
def document_json(document_id: int, session: Session = Depends(get_session)) -> dict[str, object]:
    document = session.get(IncomingDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document niet gevonden.")

    return {
        "id": document.id,
        "filename": document.original_filename,
        "project_name": document.project_name,
        "status": document.status,
        "document_type": document.document_type,
        "parser_stage": document.parser_stage or "",
        "parser_progress": document.parser_progress or 0,
        "parser_notes": document.parser_notes or "",
        "line_count": len(document.budget_lines),
        "parsed_text_available": bool(document.parsed_text),
        "source_total_amount": float(document.source_total_amount) if document.source_total_amount is not None else None,
        "source_total_source": document.source_total_source or "",
        "fields": [
            {
                "name": field.field_name,
                "value": field.field_value,
                "confidence": field.confidence,
                "source": field.source,
            }
            for field in document.fields
        ],
    }


def _decimal_or_none(value: str) -> Decimal | None:
    raw_value = str(value or "")
    cleaned = (
        raw_value.strip()
        .replace("\xa0", " ")
        .replace("€", "")
        .replace("$", "")
        .replace("EUR", "")
        .replace("eur", "")
        .replace(" ", "")
        .replace("−", "-")
    )
    cleaned = "".join(char for char in cleaned if char.isdigit() or char in {",", ".", "-"})
    if not cleaned:
        return None
    if cleaned in {"-", "--", ".", ",", "-.", "-,"}:
        return None
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "." in cleaned:
        if cleaned.count(".") > 1 or re.fullmatch(r"-?\d{1,3}(?:\.\d{3})+", cleaned):
            cleaned = cleaned.replace(".", "")
    elif "," in cleaned:
        if cleaned.count(",") > 1 or re.fullmatch(r"-?\d{1,3}(?:,\d{3})+", cleaned):
            cleaned = cleaned.replace(",", "")
        else:
            cleaned = cleaned.replace(",", ".")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _consult_lines(
    session: Session,
    query: str | None,
    status: str | None,
    date_from: str | None,
    date_to: str | None,
    priced: str | None = None,
    min_score: str | None = None,
    limit: int = 300,
) -> list[BudgetLine]:
    client_relation = aliased(Relation)
    architect_relation = aliased(Relation)
    constructor_relation = aliased(Relation)
    statement = (
        select(BudgetLine)
        .join(BudgetLine.document)
        .outerjoin(IncomingDocument.project)
        .outerjoin(client_relation, Project.client_relation_id == client_relation.id)
        .outerjoin(architect_relation, Project.architect_relation_id == architect_relation.id)
        .outerjoin(constructor_relation, Project.constructor_relation_id == constructor_relation.id)
    )
    cleaned_query = (query or "").strip()
    if cleaned_query:
        search = f"%{cleaned_query}%"
        statement = statement.where(
            or_(
                IncomingDocument.original_filename.ilike(search),
                IncomingDocument.project_name.ilike(search),
                IncomingDocument.document_type.ilike(search),
                IncomingDocument.source.ilike(search),
                Project.name.ilike(search),
                Project.project_number.ilike(search),
                Project.location.ilike(search),
                client_relation.name.ilike(search),
                architect_relation.name.ilike(search),
                constructor_relation.name.ilike(search),
                BudgetLine.omschrijving_werkzaamheden.ilike(search),
                BudgetLine.normalized_omschrijving.ilike(search),
                BudgetLine.normalized_key.ilike(search),
                BudgetLine.normalization_candidate.ilike(search),
                BudgetLine.eenheid.ilike(search),
            )
        )
    if status:
        statement = statement.where(IncomingDocument.status == status)
    if priced:
        statement = statement.where(
            or_(
                BudgetLine.eenheidsprijs.is_not(None),
                and_(
                    BudgetLine.totaal_prijs_per_regel.is_not(None),
                    BudgetLine.hoeveelheid.is_not(None),
                    BudgetLine.hoeveelheid != 0,
                ),
            )
        )
    minimum_score = _int_or_none(min_score)
    if minimum_score is not None:
        statement = statement.where(BudgetLine.confidence >= minimum_score)
    parsed_from = _date_or_none(date_from)
    if parsed_from:
        statement = statement.where(IncomingDocument.created_at >= parsed_from)
    parsed_to = _date_or_none(date_to)
    if parsed_to:
        statement = statement.where(IncomingDocument.created_at < parsed_to + timedelta(days=1))

    return session.scalars(
        statement.order_by(IncomingDocument.created_at.desc(), BudgetLine.line_number).limit(limit)
    ).unique().all()


def _reference_lines(session: Session, query: str | None, limit: int = 300) -> list[ReferenceLine]:
    statement = select(ReferenceLine).join(ReferenceLine.dataset)
    cleaned_query = (query or "").strip()
    if cleaned_query:
        search = f"%{cleaned_query}%"
        statement = statement.where(
            or_(
                ReferenceDataset.name.ilike(search),
                ReferenceDataset.source.ilike(search),
                ReferenceLine.project_name.ilike(search),
                ReferenceLine.relation_name.ilike(search),
                ReferenceLine.omschrijving_werkzaamheden.ilike(search),
                ReferenceLine.normalized_omschrijving.ilike(search),
                ReferenceLine.normalized_key.ilike(search),
                ReferenceLine.eenheid.ilike(search),
                ReferenceLine.phase.ilike(search),
                ReferenceLine.period.ilike(search),
                ReferenceLine.project_sheet_name.ilike(search),
            )
        )
    return session.scalars(
        statement.order_by(ReferenceDataset.created_at.desc(), ReferenceLine.line_number).limit(limit)
    ).unique().all()


def _reference_index_sections(session: Session, query: str | None, limit: int = 6000) -> list[dict[str, object]]:
    category_keys = {key for key, _label in REFERENCE_CATEGORY_COLUMNS}
    index_lookup = _price_index_lookup(session)
    statement = (
        select(ReferenceLine)
        .join(ReferenceLine.dataset)
        .where(ReferenceLine.regel_type == REFERENCE_INDEX_TYPE)
    )
    cleaned_query = (query or "").strip()
    if cleaned_query:
        search = f"%{cleaned_query}%"
        statement = statement.where(
            or_(
                ReferenceDataset.name.ilike(search),
                ReferenceDataset.original_filename.ilike(search),
                ReferenceLine.project_name.ilike(search),
                ReferenceLine.relation_name.ilike(search),
                ReferenceLine.hoofdstuk_omschrijving.ilike(search),
                ReferenceLine.project_sheet_name.ilike(search),
                ReferenceLine.phase.ilike(search),
                ReferenceLine.period.ilike(search),
            )
        )

    lines = session.scalars(
        statement.order_by(
            ReferenceLine.hoofdstuk_omschrijving,
            ReferenceLine.source_row,
            ReferenceLine.project_name,
            ReferenceLine.line_number,
        ).limit(limit)
    ).unique().all()
    grouped: dict[tuple[object, ...], dict[str, object]] = {}
    section_order: dict[str, int] = {}

    for line in lines:
        category_key = (line.post_code or line.hoofdstuk_code or "").lower()
        if category_key not in category_keys:
            continue
        section = (line.hoofdstuk_omschrijving or "Overig").strip() or "Overig"
        if section not in section_order:
            section_order[section] = len(section_order)
        raw_meta = _reference_raw_meta(line.raw_text)
        source_row = line.source_row or _int_or_none(raw_meta.get("rij"))
        phase = line.phase or raw_meta.get("fase") or ""
        period = line.period or raw_meta.get("periode") or ""
        stored_bdb_indexering = line.bdb_indexering if line.bdb_indexering is not None else line.norm_arbeid
        bdb_indexering = _calculated_bdb_indexering(index_lookup, period, stored_bdb_indexering)
        key = (
            line.dataset_id,
            section,
            line.project_name or "",
            line.relation_name or "",
            line.document_date.strftime("%Y-%m-%d") if line.document_date else "",
            phase,
            period,
            str(bdb_indexering or ""),
        )
        if key not in grouped:
            modal_key = str(source_row or abs(hash(key))).replace("-", "m")
            grouped[key] = {
                "dataset_id": line.dataset_id,
                "dataset_name": line.dataset.name,
                "source_filename": line.dataset.original_filename or "",
                "section": section,
                "source_row": source_row or line.line_number,
                "source_row_label": source_row or "",
                "project_name": line.project_name or "",
                "variant": line.relation_name or "",
                "phase": phase,
                "period": period,
                "bdb_indexering": bdb_indexering,
                "index_reference": _reference_index_reference(index_lookup, period, bdb_indexering),
                "project_sheet_name": line.project_sheet_name or "",
                "document_date": line.document_date,
                "date_input": line.document_date.strftime("%Y-%m-%d") if line.document_date else "",
                "date_label": line.document_date.strftime("%d-%m-%Y") if line.document_date else "",
                "categories": {},
                "category_ids": {},
                "category_sources": {},
                "source_rows": [],
                "category_count": 0,
                "total": Decimal("0"),
                "modal_id": f"kengetal-{line.dataset_id}-{modal_key}",
            }
        row_data = grouped[key]
        value = line.eenheidsprijs if line.eenheidsprijs is not None else line.totaal_prijs_per_regel
        row_data["categories"][category_key] = value
        row_data["category_ids"][category_key] = line.id
        row_data["category_sources"][category_key] = _reference_line_source_label(line, raw_meta, category_key)
        row_data["category_count"] = int(row_data["category_count"]) + 1
        if source_row and source_row not in row_data["source_rows"]:
            row_data["source_rows"].append(source_row)
            row_data["source_row_label"] = ", ".join(str(row) for row in row_data["source_rows"][:4])
            if len(row_data["source_rows"]) > 4:
                row_data["source_row_label"] += "..."
        if value is not None:
            row_data["total"] = Decimal(str(row_data["total"])) + Decimal(str(value))
        if not row_data["phase"] and phase:
            row_data["phase"] = phase
        if not row_data["period"] and period:
            row_data["period"] = period
        if row_data["bdb_indexering"] is None and bdb_indexering is not None:
            row_data["bdb_indexering"] = bdb_indexering
        if not row_data["project_sheet_name"] and line.project_sheet_name:
            row_data["project_sheet_name"] = line.project_sheet_name

    sections: dict[str, list[dict[str, object]]] = {}
    for row_data in grouped.values():
        sections.setdefault(str(row_data["section"]), []).append(row_data)

    preferred_order = {
        "bedrijfshallen met kantoor": 0,
        "kantoor": 1,
        "woningen": 2,
    }
    result: list[dict[str, object]] = []
    for section, rows in sections.items():
        rows.sort(key=lambda item: (int(item["source_row"] or 0), str(item["project_name"]).lower()))
        result.append(
            {
                "section": section,
                "rows": rows,
                "row_count": len(rows),
                "category_count": sum(int(row["category_count"]) for row in rows),
                "sort_key": (preferred_order.get(section.lower(), 50), section_order.get(section, 999), section.lower()),
            }
        )
    result.sort(key=lambda item: item["sort_key"])
    return result


def _reference_index_summary(session: Session, sections: list[dict[str, object]]) -> dict[str, object]:
    project_count = sum(int(section["row_count"]) for section in sections)
    category_line_count = sum(int(section["category_count"]) for section in sections)
    latest_index = session.scalars(select(PriceIndexValue).order_by(PriceIndexValue.effective_date.desc()).limit(1)).first()
    index_values = session.scalars(select(PriceIndexValue).order_by(PriceIndexValue.effective_date.desc()).limit(8)).all()
    index_label = ""
    index_value = None
    if latest_index is not None:
        index_label = latest_index.notes or latest_index.effective_date.strftime("%Y-%m")
        index_value = latest_index.index_value
    return {
        "project_count": project_count,
        "category_line_count": category_line_count,
        "section_count": len(sections),
        "latest_index_label": index_label,
        "latest_index_value": index_value,
        "index_values": [
            {
                "period": value.notes or value.effective_date.strftime("%Y-%m"),
                "value": value.index_value,
            }
            for value in index_values
        ],
    }


def _price_index_lookup(session: Session) -> dict[str, PriceIndexValue]:
    values = session.scalars(select(PriceIndexValue).order_by(PriceIndexValue.effective_date.desc())).all()
    lookup: dict[str, PriceIndexValue] = {}
    if values:
        lookup["__latest__"] = values[0]
    for value in values:
        keys = {
            value.notes or "",
            f"{value.effective_date.year}-Q{((value.effective_date.month - 1) // 3) + 1}",
        }
        for key in keys:
            cleaned = _period_key(key)
            if cleaned:
                lookup[cleaned] = value
    return lookup


def _period_key(value: str | None) -> str:
    return re.sub(r"[^0-9q]+", "", (value or "").lower())


def _calculated_bdb_indexering(
    index_lookup: dict[str, PriceIndexValue],
    period: str | None,
    stored_bdb_indexering: Decimal | None,
) -> Decimal | None:
    index_value = index_lookup.get(_period_key(period))
    latest_index = index_lookup.get("__latest__")
    if index_value is None or latest_index is None or not index_value.index_value:
        return stored_bdb_indexering
    try:
        return (
            Decimal(str(latest_index.index_value)) / Decimal(str(index_value.index_value))
        ).quantize(Decimal("0.0001"))
    except (InvalidOperation, ZeroDivisionError):
        return stored_bdb_indexering


def _reference_index_reference(
    index_lookup: dict[str, PriceIndexValue],
    period: str | None,
    bdb_indexering: Decimal | None,
) -> str:
    if not period and bdb_indexering is None:
        return ""
    index_value = index_lookup.get(_period_key(period))
    latest_index = index_lookup.get("__latest__")
    parts: list[str] = []
    if index_value is not None:
        if latest_index is not None and latest_index.id != index_value.id and index_value.index_value:
            factor = Decimal(str(latest_index.index_value)) / Decimal(str(index_value.index_value))
            parts.append(
                f"{latest_index.notes or latest_index.effective_date.strftime('%Y-%m')}: "
                f"{amount(latest_index.index_value)} / {index_value.notes or period}: "
                f"{amount(index_value.index_value)} = {amount(factor)}"
            )
        else:
            parts.append(f"{index_value.notes or period}: {amount(index_value.index_value)}")
        if index_value.source_reference:
            parts.append(index_value.source_reference)
    if bdb_indexering is not None:
        parts.append(f"BDB factor {amount(bdb_indexering)}")
    return " | ".join(parts)


def _reference_line_source_label(line: ReferenceLine, raw_meta: dict[str, str], category_key: str) -> str:
    parts = [line.dataset.original_filename or line.dataset.name]
    if line.project_sheet_name:
        parts.append(f"tabblad {line.project_sheet_name}")
    source_row = line.source_row or _int_or_none(raw_meta.get("rij")) or line.line_number
    if source_row:
        parts.append(f"rij {source_row}")
    if category_key:
        parts.append(f"kolom {category_key}")
    return " | ".join(parts)


def _reference_raw_meta(raw_text: str | None) -> dict[str, str]:
    meta: dict[str, str] = {}
    for part in (raw_text or "").split("|"):
        if ":" not in part:
            continue
        key, value = part.split(":", 1)
        key = key.strip().lower()
        value = value.strip()
        if key in {"rij", "fase", "periode", "categorie"} and value:
            meta[key] = value
    return meta


def _normalization_candidate_lines(session: Session, limit: int = 120) -> list[BudgetLine]:
    return session.scalars(
        select(BudgetLine)
        .join(BudgetLine.document)
        .where(
            BudgetLine.normalization_score < 100,
            or_(
                BudgetLine.normalization_candidate.is_not(None),
                and_(
                    BudgetLine.normalization_method.in_(["raw", "reference"]),
                    BudgetLine.eenheidsprijs.is_not(None),
                    BudgetLine.omschrijving_werkzaamheden != "",
                ),
            ),
        )
        .order_by(BudgetLine.normalization_candidate.is_(None), BudgetLine.normalization_score.desc(), BudgetLine.id.desc())
        .limit(limit)
    ).unique().all()


def _budget_line_groups(lines: list[BudgetLine]) -> list[dict[str, object]]:
    groups: dict[str, dict[str, object]] = {}
    for line in lines:
        description = (line.omschrijving_werkzaamheden or "").strip()
        score = int(line.normalization_score or 0)
        method = line.normalization_method or "raw"
        if method == "noise" or is_noise_line(description):
            label = "Lage kwaliteit / opschonen"
            key = "noise"
        else:
            label = (
                line.normalization_candidate
                or line.normalized_omschrijving
                or description
                or "Niet genormaliseerd"
            ).strip()
            key = normalization_key(label) or label.lower()
        group = groups.setdefault(
            key,
            {
                "label": label,
                "lines": [],
                "best_score": 0,
                "methods": set(),
                "priced_count": 0,
            },
        )
        group["lines"].append(line)
        group["best_score"] = max(int(group["best_score"]), score)
        group["methods"].add(method)
        if calculated_unit_price(line) is not None or line.totaal_prijs_per_regel is not None:
            group["priced_count"] = int(group["priced_count"]) + 1

    grouped: list[dict[str, object]] = []
    for group in groups.values():
        group_lines = sorted(
            group["lines"],
            key=lambda item: (
                -(item.normalization_score or 0),
                1 if (item.normalization_method == "noise" or is_noise_line(item.omschrijving_werkzaamheden)) else 0,
                item.line_number,
                item.id,
            ),
        )
        methods = sorted(str(method) for method in group["methods"] if method)
        grouped.append(
            {
                "label": group["label"],
                "lines": group_lines,
                "best_score": int(group["best_score"]),
                "method_label": ", ".join(methods) if methods else "raw",
                "count": len(group_lines),
                "priced_count": int(group["priced_count"]),
            }
        )
    return sorted(
        grouped,
        key=lambda item: (
            1 if str(item["label"]).startswith("Lage kwaliteit") else 0,
            -int(item["best_score"]),
            -int(item["priced_count"]),
            str(item["label"]).lower(),
        ),
    )


def _document_total_context(document: IncomingDocument) -> dict[str, object]:
    line_total = _budget_line_total(document.budget_lines)
    quantity_total = _budget_quantity_total(document.budget_lines)
    source_total, source_label = _source_total_from_document(document)
    difference = None
    difference_abs = None
    match_class = "missing"
    match_label = "Origineel totaal niet herkend"
    if source_total is not None:
        difference = line_total - source_total
        difference_abs = abs(difference)
        if difference_abs <= Decimal("1.00"):
            match_class = "ok"
            match_label = "Sluit aan"
        elif difference_abs <= max(Decimal("25.00"), abs(source_total) * Decimal("0.0025")):
            match_class = "warn"
            match_label = "Klein verschil"
        else:
            match_class = "error"
            match_label = "Controleer verschil"
    return {
        "line_total": line_total,
        "quantity_total": quantity_total,
        "source_total": source_total,
        "source_label": source_label,
        "difference": difference,
        "difference_abs": difference_abs,
        "match_class": match_class,
        "match_label": match_label,
    }


def _budget_line_total(lines: list[BudgetLine]) -> Decimal:
    usable_lines = [
        line
        for line in lines
        if line.regel_type not in {"hoofdstuk", "post", "subtotaal", "totaal"}
        and not is_noise_line(line.omschrijving_werkzaamheden)
        and calculated_line_total(line) is not None
    ]
    if not usable_lines:
        usable_lines = [line for line in lines if calculated_line_total(line) is not None]
    total = Decimal("0")
    for line in usable_lines:
        total += Decimal(str(calculated_line_total(line) or 0))
    return total


def _budget_quantity_total(lines: list[BudgetLine]) -> Decimal:
    total = Decimal("0")
    for line in lines:
        if line.regel_type in {"hoofdstuk", "post", "subtotaal", "totaal"}:
            continue
        if is_noise_line(line.omschrijving_werkzaamheden):
            continue
        unit_price = calculated_unit_price(line)
        if line.hoeveelheid in {None, 0} or unit_price is None:
            continue
        try:
            total += Decimal(str(line.hoeveelheid)) * Decimal(str(unit_price))
        except InvalidOperation:
            continue
    return total


def _source_total_from_document(
    document: IncomingDocument,
    include_saved: bool = True,
) -> tuple[Decimal | None, str | None]:
    if include_saved and document.source_total_amount is not None:
        label = document.source_total_source or "opgeslagen origineel totaal"
        return Decimal(str(document.source_total_amount)), label

    candidates: list[tuple[int, Decimal, str]] = []
    strong_keywords = (
        "eindtotaal",
        "eind totaal",
        "totaal",
        "aanneemsom",
        "inschrijfsom",
        "bouwkosten",
        "begrotingstotaal",
        "totaal begroting",
        "totaal bouwkosten",
    )

    for field in document.fields:
        haystack = f"{field.field_name} {field.field_value}".lower()
        amount_value = _last_amount_decimal(field.field_value)
        if amount_value is None:
            continue
        if any(keyword in haystack for keyword in strong_keywords):
            candidates.append((4, amount_value, f"veld: {field.field_name}"))
        elif field.field_name.lower() == "bedrag":
            candidates.append((1, amount_value, "veld: bedrag"))

    for raw_line in (document.parsed_text or "").splitlines():
        line = " ".join(raw_line.split())
        lowered = line.lower()
        if not any(keyword in lowered for keyword in strong_keywords):
            continue
        amount_value = _last_amount_decimal(line)
        if amount_value is not None:
            candidates.append((3, amount_value, "originele tekst"))

    for line in document.budget_lines:
        description = (line.omschrijving_werkzaamheden or "").lower()
        if line.totaal_prijs_per_regel is not None and (
            line.regel_type == "totaal" or any(keyword in description for keyword in strong_keywords)
        ):
            candidates.append((2, Decimal(str(line.totaal_prijs_per_regel)), "totaalregel"))

    if not candidates:
        return None, None
    candidates.sort(key=lambda item: (item[0], abs(item[1])), reverse=True)
    _, amount_value, label = candidates[0]
    return amount_value, label


def _last_amount_decimal(value: str | None) -> Decimal | None:
    matches = MONEY_TEXT_PATTERN.findall(str(value or ""))
    for match in reversed(matches):
        parsed = _decimal_or_none(match)
        if parsed is not None:
            return parsed
    return None


def _normalization_candidate_groups(lines: list[BudgetLine]) -> list[dict[str, object]]:
    groups: dict[str, dict[str, object]] = {}
    for line in lines:
        label = (
            line.normalization_candidate
            or line.normalized_omschrijving
            or line.omschrijving_werkzaamheden
            or "Nieuwe term nodig"
        ).strip()
        key = normalization_key(label) or label.lower()
        group = groups.setdefault(
            key,
            {
                "label": label,
                "lines": [],
                "best_score": 0,
                "methods": set(),
            },
        )
        group["lines"].append(line)
        group["best_score"] = max(int(group["best_score"]), int(line.normalization_score or 0))
        if line.normalization_method:
            group["methods"].add(line.normalization_method)

    grouped = []
    for group in groups.values():
        group_lines = sorted(
            group["lines"],
            key=lambda item: (-(item.normalization_score or 0), item.line_number, item.id),
        )
        methods = sorted(str(method) for method in group["methods"])
        grouped.append(
            {
                "label": group["label"],
                "lines": group_lines,
                "best_score": group["best_score"],
                "method_label": ", ".join(methods) if methods else "raw",
                "count": len(group_lines),
            }
        )
    return sorted(grouped, key=lambda item: (-int(item["best_score"]), str(item["label"]).lower()))


def _normalization_candidate_score_levels(lines: list[BudgetLine]) -> list[dict[str, object]]:
    levels = [
        {
            "key": "high",
            "label": "75 - 100",
            "title": "Sterke voorstellen",
            "description": "Waarschijnlijk goed, snel valideren.",
            "groups": [],
            "count": 0,
        },
        {
            "key": "mid",
            "label": "50 - 75",
            "title": "Twijfelgevallen",
            "description": "Controleren voordat je ze hard maakt.",
            "groups": [],
            "count": 0,
        },
        {
            "key": "low",
            "label": "0 - 50",
            "title": "Aanvullen",
            "description": "Waarschijnlijk handmatig aanvullen of negeren.",
            "groups": [],
            "count": 0,
        },
    ]
    by_key = {str(level["key"]): level for level in levels}
    for group in _normalization_candidate_groups(lines):
        score = int(group["best_score"] or 0)
        if score >= 75:
            level = by_key["high"]
        elif score >= 50:
            level = by_key["mid"]
        else:
            level = by_key["low"]
        level["groups"].append(group)
        level["count"] = int(level["count"]) + int(group["count"])
    return [level for level in levels if level["groups"]]


def _normalization_stats(session: Session) -> dict[str, int]:
    budget_total = session.scalar(select(func.count(BudgetLine.id))) or 0
    reference_total = session.scalar(select(func.count(ReferenceLine.id))) or 0
    term_count = session.scalar(select(func.count(NormalizationTerm.id))) or 0
    suggestions = session.scalar(
        select(func.count(BudgetLine.id)).where(BudgetLine.normalization_candidate.is_not(None))
    ) or 0
    hard_matches = session.scalar(
        select(func.count(BudgetLine.id)).where(BudgetLine.normalization_method == "hard")
    ) or 0
    fuzzy_matches = session.scalar(
        select(func.count(BudgetLine.id)).where(BudgetLine.normalization_method.in_(["fuzzy", "reference"]))
    ) or 0
    raw_priced = session.scalar(
        select(func.count(BudgetLine.id)).where(
            BudgetLine.normalization_method.in_(["raw", "reference"]),
            BudgetLine.eenheidsprijs.is_not(None),
            BudgetLine.normalization_score < 100,
        )
    ) or 0
    return {
        "budget_total": int(budget_total),
        "reference_total": int(reference_total),
        "term_count": int(term_count),
        "suggestions": int(suggestions) + int(raw_priced),
        "hard_matches": int(hard_matches),
        "fuzzy_matches": int(fuzzy_matches),
    }


def _add_normalization_terms(
    session: Session,
    canonical_label: str,
    aliases: str,
    category: str,
    match_type: str,
    min_score: str,
) -> list[NormalizationTerm]:
    label = canonical_label.strip()
    if not label:
        raise HTTPException(status_code=400, detail="Vul een standaardomschrijving in.")
    term_type = match_type if match_type in {"hard", "fuzzy"} else "fuzzy"
    score = max(50, min(100, _int_or_none(min_score) or 82))
    alias_values = split_aliases(aliases) or [label]
    canonical_key = normalization_key(label) or label.lower().replace(" ", "_")[:180]
    existing = {
        (term.canonical_key, normalization_key(term.alias))
        for term in session.scalars(select(NormalizationTerm)).all()
    }
    added: list[NormalizationTerm] = []
    for alias in alias_values:
        lookup = (canonical_key, normalization_key(alias))
        if lookup in existing:
            continue
        term = NormalizationTerm(
            canonical_key=canonical_key,
            canonical_label=label,
            alias=alias.strip(),
            category=category.strip() or "omschrijving",
            match_type=term_type,
            min_score=score,
            active=1,
        )
        session.add(term)
        added.append(term)
        existing.add(lookup)
    return added


def _reapply_all_normalization(session: Session) -> None:
    reference_lines = session.scalars(select(ReferenceLine)).all()
    apply_normalization(session, reference_lines)
    session.flush()

    budget_lines = session.scalars(select(BudgetLine)).all()
    for line in budget_lines:
        _normalize_line_prices(line)
    apply_normalization(session, budget_lines)
    session.commit()


def _parse_document_background(document_id: int, usage_source: str, force_openai: bool = False) -> None:
    session = SessionLocal()
    try:
        document = session.get(IncomingDocument, document_id)
        if document is None:
            return
        _set_parse_progress(session, document, 12, "Bestand voorbereiden")
        file_path = settings.upload_dir / document.stored_filename
        if not file_path.exists():
            document.status = "parse_error"
            document.parser_notes = "Parser fout: bestand niet gevonden op de server."
            document.parser_stage = "Bestand niet gevonden"
            document.parser_progress = 100
            session.commit()
            return

        _set_parse_progress(session, document, 28, "PDF lezen, OCR/OpenAI voorbereiden")
        parsed = parse_pdf(file_path, force_openai=force_openai)
        _set_parse_progress(session, document, 76, f"{len(parsed.budget_lines)} regels herkend")
        document.source = parsed.parse_method
        document.parsed_text = parsed.text
        document.parser_notes = _merge_parser_notes(document.parser_notes, _parser_note(parsed))
        document.fields = [
            ExtractedField(field_name=field.name, field_value=field.value, confidence=field.confidence)
            for field in parsed.fields
        ]
        _set_parse_progress(session, document, 88, "Begrotingsregels opslaan")
        document.budget_lines = [
            _budget_line_from_parsed(line)
            for line in parsed.budget_lines
            if not is_noise_line(line.omschrijving_werkzaamheden)
        ]
        for line in document.budget_lines:
            _normalize_line_prices(line)
        apply_normalization(session, document.budget_lines)
        if document.source_total_manual != 1:
            source_total, source_label = _source_total_from_document(document, include_saved=False)
            document.source_total_amount = source_total
            document.source_total_source = source_label
        document.status = "needs_review"
        document.parser_stage = "Klaar voor controle"
        document.parser_progress = 100
        session.commit()
        _record_openai_usage(session, usage_source, document.id, parsed.openai_usage)
    except Exception as exc:
        session.rollback()
        document = session.get(IncomingDocument, document_id)
        if document is not None:
            document.status = "parse_error"
            document.parser_notes = f"Parser fout: {str(exc)[:500]}"
            document.parser_stage = "Parser fout"
            document.parser_progress = 100
            session.commit()
    finally:
        session.close()


def _set_parse_progress(session: Session, document: IncomingDocument, progress: int, stage: str) -> None:
    document.status = "processing"
    document.parser_progress = max(0, min(100, progress))
    document.parser_stage = stage
    session.commit()


def _budget_line_from_parsed(line) -> BudgetLine:
    return BudgetLine(
        line_number=line.line_number,
        regel_type=line.regel_type,
        niveau=line.niveau,
        hoofdstuk_code=line.hoofdstuk_code,
        hoofdstuk_omschrijving=line.hoofdstuk_omschrijving,
        post_code=line.post_code,
        omschrijving_werkzaamheden=line.omschrijving_werkzaamheden,
        hoeveelheid=line.hoeveelheid,
        eenheid=line.eenheid,
        norm_arbeid=line.norm_arbeid,
        uren=line.uren,
        materiaal=line.materiaal,
        materieel=line.materieel,
        onderaannemer=line.onderaannemer,
        totaal_prijs_per_regel=line.totaal_prijs_per_regel,
        eenheidsprijs=line.eenheidsprijs,
        bron_pagina=line.bron_pagina,
        confidence=line.confidence,
        raw_text=line.raw_text,
    )


def _merge_parser_notes(existing: str | None, parsed_note: str | None) -> str | None:
    keep = []
    for part in (existing or "").split("|"):
        cleaned = part.strip()
        if not cleaned:
            continue
        lowered = cleaned.lower()
        if "parser loopt" in lowered or "upload ontvangen" in lowered or "herparse gestart" in lowered:
            continue
        keep.append(cleaned)
    if parsed_note:
        keep.append(parsed_note)
    return " | ".join(dict.fromkeys(keep)) if keep else None


def _latest_template(session: Session) -> AssessmentTemplate | None:
    return session.scalars(
        select(AssessmentTemplate)
        .where(AssessmentTemplate.status == "active")
        .order_by(AssessmentTemplate.created_at.desc())
        .limit(1)
    ).first()


def _template_for_document(
    session: Session,
    document: IncomingDocument,
    template_id: str | int | None = None,
) -> AssessmentTemplate | None:
    explicit_template_id = _int_or_none(template_id) if template_id is not None else None
    if explicit_template_id:
        template = session.get(AssessmentTemplate, explicit_template_id)
        if template is not None:
            return template
    for part in (document.parser_notes or "").split("|"):
        key, _, value = part.strip().partition("=")
        if key == "template_id":
            template_id = _int_or_none(value)
            if template_id:
                template = session.get(AssessmentTemplate, template_id)
                if template is not None:
                    return template
    return _latest_template(session)


def _run_job(session: Session, job: ScheduledJob) -> int:
    if job.job_type == "index_sync":
        if job.target_id:
            series = session.get(PriceIndexSeries, job.target_id)
            if series is None:
                raise ValueError("Indexreeks niet gevonden.")
            return sync_price_index_series(session, series)
        total = 0
        for series in session.scalars(select(PriceIndexSeries).where(PriceIndexSeries.api_url.is_not(None))).all():
            total += sync_price_index_series(session, series)
        return total
    raise ValueError("Onbekend jobtype.")


def _add_budget_lines_to_references(
    session: Session,
    document: IncomingDocument,
    lines: list[BudgetLine],
) -> int:
    dataset = _assessment_reference_dataset(session)
    next_line_number = (
        session.scalar(
            select(func.max(ReferenceLine.line_number)).where(ReferenceLine.dataset_id == dataset.id)
        )
        or 0
    ) + 1
    added_lines: list[ReferenceLine] = []
    project = document.project
    project_name = project.name if project else document.project_name
    relation_name = project.client.name if project and project.client else None

    for line in lines:
        description = (line.omschrijving_werkzaamheden or "").strip()
        unit_price = calculated_unit_price(line)
        if not description or unit_price is None or is_noise_line(description):
            continue
        marker = f"budget_line:{line.id}"
        exists = session.scalar(
            select(ReferenceLine.id)
            .where(ReferenceLine.dataset_id == dataset.id)
            .where(ReferenceLine.raw_text == marker)
            .limit(1)
        )
        if exists:
            continue
        reference_line = ReferenceLine(
            dataset_id=dataset.id,
            line_number=next_line_number,
            regel_type=line.regel_type,
            niveau=line.niveau,
            hoofdstuk_code=line.hoofdstuk_code,
            hoofdstuk_omschrijving=line.hoofdstuk_omschrijving,
            post_code=line.post_code,
            project_name=project_name,
            relation_name=relation_name,
            document_date=document.created_at,
            omschrijving_werkzaamheden=description,
            hoeveelheid=line.hoeveelheid,
            eenheid=line.eenheid,
            norm_arbeid=line.norm_arbeid,
            uren=line.uren,
            materiaal=line.materiaal,
            materieel=line.materieel,
            onderaannemer=line.onderaannemer,
            totaal_prijs_per_regel=calculated_line_total(line),
            eenheidsprijs=unit_price,
            bron_pagina=line.bron_pagina,
            normalized_key=line.normalized_key,
            normalized_omschrijving=line.normalized_omschrijving,
            normalization_method=line.normalization_method,
            normalization_score=line.normalization_score,
            normalization_candidate=line.normalization_candidate,
            confidence=max(line.confidence or 0, 80),
            raw_text=marker,
        )
        session.add(reference_line)
        added_lines.append(reference_line)
        next_line_number += 1

    if added_lines:
        apply_normalization(session, added_lines)
    return len(added_lines)


def _assessment_reference_dataset(session: Session) -> ReferenceDataset:
    dataset_name = "Beoordelingen - gevalideerde kengetallen"
    dataset = session.scalar(
        select(ReferenceDataset)
        .where(ReferenceDataset.name == dataset_name)
        .where(ReferenceDataset.source == "beoordeling")
        .limit(1)
    )
    if dataset is not None:
        return dataset
    dataset = ReferenceDataset(
        name=dataset_name,
        source="beoordeling",
        notes="Regels die vanuit begrotingsbeoordelingen als kengetal zijn toegevoegd.",
        status="active",
    )
    session.add(dataset)
    session.flush()
    return dataset


def _date_or_none(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return None


def _period_to_date(value: str | None) -> datetime | None:
    raw_value = (value or "").strip()
    if not raw_value:
        return None
    quarter_match = re.fullmatch(r"(\d{4})[-\s]?Q([1-4])", raw_value, flags=re.IGNORECASE)
    if quarter_match:
        year = int(quarter_match.group(1))
        month = (int(quarter_match.group(2)) - 1) * 3 + 1
        return datetime(year, month, 1)
    month_match = re.fullmatch(r"(\d{4})[-/](\d{1,2})", raw_value)
    if month_match:
        year = int(month_match.group(1))
        month = int(month_match.group(2))
        if 1 <= month <= 12:
            return datetime(year, month, 1)
    return _date_or_none(raw_value)


def _form_value(form, key: str, index: int) -> str:
    values = form.getlist(key)
    if index >= len(values):
        return ""
    return str(values[index] or "")


def _normalize_document_line_prices(document: IncomingDocument) -> bool:
    changed = False
    for line in document.budget_lines:
        changed = _normalize_line_prices(line) or changed
    return changed


def _normalize_line_prices(line: BudgetLine) -> bool:
    changed = False
    component_total = _price_component_total(line)
    if line.eenheidsprijs is None and component_total is not None:
        line.eenheidsprijs = component_total
        changed = True
    if line.totaal_prijs_per_regel is not None and line.hoeveelheid not in {None, 0}:
        try:
            unit_price = line.totaal_prijs_per_regel / line.hoeveelheid
            if line.eenheidsprijs != unit_price:
                line.eenheidsprijs = unit_price
                changed = True
        except (InvalidOperation, ZeroDivisionError):
            pass
    elif line.totaal_prijs_per_regel is None and line.eenheidsprijs is not None:
        if line.hoeveelheid not in {None, 0}:
            try:
                line.totaal_prijs_per_regel = Decimal(str(line.hoeveelheid)) * Decimal(str(line.eenheidsprijs))
                changed = True
            except InvalidOperation:
                pass
        else:
            line.totaal_prijs_per_regel = line.eenheidsprijs
            changed = True
    return changed


def _price_component_total(line: BudgetLine) -> Decimal | None:
    total = Decimal("0")
    has_component = False
    for value in (line.materiaal, line.materieel, line.onderaannemer):
        if value is None:
            continue
        total += Decimal(str(value))
        has_component = True
    return total if has_component else None


def _int_or_none(value: str | int | None) -> int | None:
    try:
        if value is None or str(value).strip() == "":
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def _flash_message(request: Request) -> str | None:
    messages = {
        "voorstel_gevalideerd": "Voorstel gevalideerd en toegevoegd aan het basiswoordenboek.",
        "term_toegevoegd": "Nieuwe normalisatieterm opgeslagen en opnieuw toegepast.",
        "normalisatie_bijgewerkt": "Alle regels zijn opnieuw genormaliseerd.",
        "origineel_totaal_opgeslagen": "Origineel inputtotaal opgeslagen voor deze beoordeling.",
        "regels_opgeslagen": "Begrotingsregels opgeslagen en opnieuw genormaliseerd.",
        "regels_verwijderd": "Geselecteerde begrotingsregels verwijderd.",
        "regels_gevalideerd": "Begrotingsregels gevalideerd en document gemarkeerd als verwerkt.",
        "kengetallen_toegevoegd": "Geselecteerde regels zijn toegevoegd aan de kengetallen-database.",
        "geen_kengetallen_geselecteerd": "Geen bruikbare geselecteerde regels met eenheidsprijs gevonden.",
        "kengetal_project_opgeslagen": "Kengetalproject opgeslagen en opnieuw genormaliseerd.",
        "index_opgeslagen": "Indexregel opgeslagen.",
        "index_ongeldig": "Indexwaarde is niet geldig.",
        "index_periode_ongeldig": "Periode is niet geldig.",
    }
    return messages.get(request.query_params.get("notice", ""))


def request_url_for_document(document_id: int, archived: bool = False) -> str:
    return "/" if archived else f"/documents/{document_id}"


def _safe_return(value: str | None) -> str | None:
    if not value:
        return None
    cleaned = value.strip()
    if cleaned.startswith("/") and not cleaned.startswith("//"):
        return cleaned
    return None


def _excel_filename(filename: str, template_suffix: str = ".xlsx") -> str:
    cleaned = re.sub(r'[^A-Za-z0-9_. -]+', "_", (filename or "").strip()) or "export.xlsx"
    suffix = ".xlsm" if template_suffix.lower() == ".xlsm" else ".xlsx"
    stem = re.sub(r"\.(xlsx|xlsm)$", "", cleaned, flags=re.IGNORECASE)
    return f"{stem}{suffix}"


def _parser_note(parsed) -> str | None:
    parts = [f"Parse methode: {parsed.parse_method}", f"zekerheid: {parsed.confidence}%"]
    if parsed.openai_usage:
        parts.append(f"OpenAI tokens: {parsed.openai_usage.get('total_tokens', 0)}")
    if parsed.notes:
        parts.append(parsed.notes)
    return " | ".join(parts)


def _assessment_note(template_id: str, reference_dataset_id: str) -> str:
    parts = ["Module: begroting beoordeling"]
    if template_id:
        parts.append(f"template_id={template_id}")
    if reference_dataset_id:
        parts.append(f"kengetallen_dataset_id={reference_dataset_id}")
    return " | ".join(parts)


def _record_openai_usage(session: Session, source: str, source_id: int, usage: dict | None) -> None:
    if not usage:
        return
    model = os.getenv("OPENAI_MODEL", "gpt-4.1-mini").strip() or "gpt-4.1-mini"
    event = OpenAIUsageEvent(
        source=source,
        source_id=source_id,
        model=model,
        input_tokens=int(usage.get("input_tokens") or 0),
        cached_input_tokens=int(usage.get("cached_input_tokens") or 0),
        output_tokens=int(usage.get("output_tokens") or 0),
        total_tokens=int(usage.get("total_tokens") or 0),
        estimated_cost_usd=_estimate_openai_cost(model, usage),
    )
    session.add(event)
    session.commit()


def _status_context(session: Session) -> dict[str, object]:
    db_status = _database_status()
    openai_usage = _openai_usage_summary(session)
    server_info = _server_info(db_status)
    return {
        "db_status": db_status,
        "openai_usage": openai_usage,
        "server_info": server_info,
    }


def _database_status() -> dict[str, object]:
    started = time.perf_counter()
    try:
        with engine.connect() as connection:
            connection.execute(text("SELECT 1"))
            database_name = connection.execute(text("SELECT current_database()")).scalar_one_or_none()
            try:
                db_size = connection.execute(text("SELECT pg_size_pretty(pg_database_size(current_database()))")).scalar_one_or_none()
            except Exception:
                db_size = "-"
        return {
            "connected": True,
            "label": "connected",
            "database": database_name or "database",
            "latency_ms": round((time.perf_counter() - started) * 1000),
            "size": db_size or "-",
        }
    except Exception as exc:
        return {
            "connected": False,
            "label": "offline",
            "database": "database",
            "latency_ms": None,
            "size": "-",
            "error": str(exc)[:120],
        }


def _openai_usage_summary(session: Session) -> dict[str, object]:
    month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    total_requests = session.scalar(select(func.count(OpenAIUsageEvent.id))) or 0
    month_requests = session.scalar(
        select(func.count(OpenAIUsageEvent.id)).where(OpenAIUsageEvent.created_at >= month_start)
    ) or 0
    month_tokens = session.scalar(
        select(func.coalesce(func.sum(OpenAIUsageEvent.total_tokens), 0)).where(OpenAIUsageEvent.created_at >= month_start)
    ) or 0
    month_cost = session.scalar(
        select(func.coalesce(func.sum(OpenAIUsageEvent.estimated_cost_usd), 0)).where(OpenAIUsageEvent.created_at >= month_start)
    ) or Decimal("0")
    return {
        "enabled": bool(os.getenv("OPENAI_API_KEY")),
        "month_requests": int(month_requests),
        "total_requests": int(total_requests),
        "month_tokens": int(month_tokens),
        "month_cost_usd": Decimal(str(month_cost or 0)),
        "model": os.getenv("OPENAI_MODEL", "gpt-4.1-mini"),
    }


def _server_info(db_status: dict[str, object]) -> dict[str, object]:
    return {
        "uptime": _format_duration(int(time.time() - APP_STARTED_AT)),
        "load": _load_percent(),
        "memory": _memory_info(),
        "database": db_status,
        "ocr_enabled": os.getenv("OCR_ENABLED", "true").strip().lower() not in {"0", "false", "nee", "no"},
        "openai_enabled": bool(os.getenv("OPENAI_API_KEY")),
    }


def _load_percent() -> dict[str, object]:
    try:
        load_1 = os.getloadavg()[0]
        cores = os.cpu_count() or 1
        return {"percent": min(999, round((load_1 / cores) * 100, 1)), "label": f"load {load_1:.2f} - {cores} cores"}
    except Exception:
        return {"percent": 0, "label": "niet beschikbaar"}


def _memory_info() -> dict[str, object]:
    try:
        meminfo = {}
        for line in Path("/proc/meminfo").read_text().splitlines():
            key, value = line.split(":", 1)
            meminfo[key] = int(value.strip().split()[0])
        total = meminfo.get("MemTotal", 0)
        available = meminfo.get("MemAvailable", 0)
        used = max(total - available, 0)
        percent = round((used / total) * 100, 1) if total else 0
        return {"percent": percent, "label": f"{used / 1024 / 1024:.1f} GB / {total / 1024 / 1024:.1f} GB"}
    except Exception:
        return {"percent": 0, "label": "niet beschikbaar"}


def _format_duration(seconds: int) -> str:
    delta = timedelta(seconds=max(seconds, 0))
    days = delta.days
    hours, remainder = divmod(delta.seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    if days:
        return f"{days}d {hours}u"
    if hours:
        return f"{hours}u {minutes}m"
    return f"{minutes}m"


def _estimate_openai_cost(model: str, usage: dict) -> Decimal:
    pricing = {
        "gpt-4.1-mini": {"input": Decimal("0.25"), "cached_input": Decimal("0.025"), "output": Decimal("2.00")},
        "gpt-4.1": {"input": Decimal("2.00"), "cached_input": Decimal("0.50"), "output": Decimal("8.00")},
        "gpt-5-mini": {"input": Decimal("0.25"), "cached_input": Decimal("0.025"), "output": Decimal("2.00")},
        "gpt-5-nano": {"input": Decimal("0.05"), "cached_input": Decimal("0.005"), "output": Decimal("0.40")},
    }.get(model, {"input": Decimal("0.25"), "cached_input": Decimal("0.025"), "output": Decimal("2.00")})
    input_tokens = max(int(usage.get("input_tokens") or 0) - int(usage.get("cached_input_tokens") or 0), 0)
    cached_tokens = int(usage.get("cached_input_tokens") or 0)
    output_tokens = int(usage.get("output_tokens") or 0)
    cost = (
        Decimal(input_tokens) * pricing["input"]
        + Decimal(cached_tokens) * pricing["cached_input"]
        + Decimal(output_tokens) * pricing["output"]
    ) / Decimal("1000000")
    return cost.quantize(Decimal("0.0001"))
