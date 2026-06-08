from __future__ import annotations

from copy import copy
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.worksheet.worksheet import Worksheet

from .models import BudgetLine, IncomingDocument


FIELD_ALIASES = {
    "hoofdstuk_code": {"hoofdstuk", "hfd", "hoofdstukcode"},
    "post_code": {"post", "postnr", "postnummer", "code"},
    "omschrijving_werkzaamheden": {"omschrijving", "omschrijving werkzaamheden", "omschrijving/ werkzaamheden", "werkzaamheden"},
    "hoeveelheid": {"hvh", "hoeveelheid", "aantal"},
    "eenheid": {"ehd", "eenheid"},
    "norm_arbeid": {"norm", "norm arbeid", "norm/ arbeid"},
    "uren": {"uren", "uur"},
    "materiaal": {"materiaal"},
    "materieel": {"materieel"},
    "onderaannemer": {"o.a.", "oa", "onderaannemer"},
    "eenheidsprijs": {"eenheidsprijs", "ehprijs", "eindprijs", "prijs per eenheid"},
    "totaal_prijs_per_regel": {"totaal", "totaalprijs", "totaal prijs per regel"},
}

MIN_HEADER_SCORE = 5

ColumnMapping = dict[str, list[int]]


def fill_screening_template(
    document: IncomingDocument,
    template_path: Path,
    target_sheet: str | None = None,
    logo_path: Path | None = None,
) -> BytesIO:
    workbook = load_workbook(template_path, keep_vba=template_path.suffix.lower() == ".xlsm")
    sheet = workbook[target_sheet] if target_sheet and target_sheet in workbook.sheetnames else _best_sheet(workbook.worksheets)
    header_row, mapping = _find_header_row(sheet)
    if not mapping:
        raise ValueError("Geen herkenbare controlemodel-kolommen gevonden in het template.")

    start_row = header_row + 1
    existing_end = _existing_body_end(sheet, start_row, mapping)
    needed_rows = max(len(document.budget_lines), 1)
    available_rows = max(existing_end - start_row + 1, 1)
    sample_row = start_row

    if needed_rows > available_rows:
        sheet.insert_rows(existing_end + 1, needed_rows - available_rows)
        for row in range(existing_end + 1, existing_end + 1 + needed_rows - available_rows):
            _copy_row_style(sheet, sample_row, row)

    _clear_body(sheet, start_row, start_row + max(available_rows, needed_rows) - 1, mapping)
    for offset, line in enumerate(document.budget_lines):
        row = start_row + offset
        _write_line(sheet, row, mapping, line)
        if line.regel_type in {"hoofdstuk", "post"}:
            _mark_structure_row(sheet, row, mapping, line)

    _write_document_header(sheet, document, logo_path)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _best_sheet(sheets: list[Worksheet]) -> Worksheet:
    for sheet in sheets:
        if sheet.title.strip().lower() == "screening":
            header_row, mapping = _find_header_row(sheet)
            if mapping:
                return sheet
    scored = []
    for sheet in sheets:
        header_row, mapping = _find_header_row(sheet)
        scored.append((len(mapping), -header_row if header_row else -9999, sheet))
    scored.sort(key=lambda item: item[:2], reverse=True)
    return scored[0][2]


def _find_header_row(sheet: Worksheet) -> tuple[int, ColumnMapping]:
    best_row = 0
    best_score = 0
    best_mapping: ColumnMapping = {}
    max_scan = min(sheet.max_row, 80)
    for row in range(1, max_scan + 1):
        values = [_normalize(sheet.cell(row=row, column=column).value) for column in range(1, sheet.max_column + 1)]
        mapping: ColumnMapping = {}
        for field, aliases in FIELD_ALIASES.items():
            for index, value in enumerate(values, start=1):
                if value in aliases:
                    mapping.setdefault(field, []).append(index)
        score = len(mapping) + sum(max(0, len(columns) - 1) for columns in mapping.values())
        if "omschrijving_werkzaamheden" in mapping and score > best_score:
            best_row = row
            best_score = score
            best_mapping = mapping
    if len(best_mapping) < MIN_HEADER_SCORE:
        return 0, {}
    return best_row, best_mapping


def _existing_body_end(sheet: Worksheet, start_row: int, mapping: ColumnMapping) -> int:
    columns = _mapped_columns(mapping)
    last_seen = start_row
    empty_streak = 0
    for row in range(start_row, min(sheet.max_row, start_row + 500) + 1):
        has_value = any(sheet.cell(row=row, column=column).value not in {None, ""} for column in columns)
        if has_value:
            last_seen = row
            empty_streak = 0
        else:
            empty_streak += 1
        if empty_streak >= 5:
            break
    return max(last_seen, start_row)


def _clear_body(sheet: Worksheet, start_row: int, end_row: int, mapping: ColumnMapping) -> None:
    for row in range(start_row, end_row + 1):
        for column in _mapped_columns(mapping):
            sheet.cell(row=row, column=column).value = None


def _copy_row_style(sheet: Worksheet, source_row: int, target_row: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(row=source_row, column=column)
        target = sheet.cell(row=target_row, column=column)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def _write_line(sheet: Worksheet, row: int, mapping: ColumnMapping, line: BudgetLine) -> None:
    values = {
        "hoofdstuk_code": line.hoofdstuk_code or line.post_code,
        "post_code": line.post_code or line.hoofdstuk_code,
        "omschrijving_werkzaamheden": line.omschrijving_werkzaamheden,
        "hoeveelheid": _amount(line.hoeveelheid),
        "eenheid": line.eenheid,
        "norm_arbeid": _amount(line.norm_arbeid),
        "uren": _amount(line.uren),
        "materiaal": line.materiaal,
        "materieel": line.materieel,
        "onderaannemer": line.onderaannemer,
        "eenheidsprijs": _unit_price(line),
        "totaal_prijs_per_regel": line.totaal_prijs_per_regel,
    }
    for field, value in values.items():
        for column in mapping.get(field, []):
            sheet.cell(row=row, column=column).value = value


def _mark_structure_row(sheet: Worksheet, row: int, mapping: ColumnMapping, line: BudgetLine) -> None:
    description_columns = mapping.get("omschrijving_werkzaamheden", [])
    for description_column in description_columns:
        prefix = line.hoofdstuk_code or line.post_code
        if prefix and not str(sheet.cell(row=row, column=description_column).value or "").startswith(prefix):
            sheet.cell(row=row, column=description_column).value = f"{prefix} {line.omschrijving_werkzaamheden}"
    for column in _mapped_columns(mapping):
        cell = sheet.cell(row=row, column=column)
        font = copy(cell.font)
        font.bold = True
        cell.font = font


def _write_document_header(sheet: Worksheet, document: IncomingDocument, logo_path: Path | None) -> None:
    if sheet.title.strip().lower() != "screening":
        return
    if document.project_name:
        _write_if_empty(sheet, "AK4", document.project_name)
    if document.original_filename:
        _write_if_empty(sheet, "AK5", document.original_filename)
    if document.created_at:
        _write_if_empty(sheet, "AN7", document.created_at.strftime("%d-%m-%Y"))
    if document.source_total_amount is not None:
        _write_if_empty(sheet, "AN8", document.source_total_amount)
    if logo_path and logo_path.exists() and not getattr(sheet, "_images", []):
        try:
            image = ExcelImage(str(logo_path))
            image.width = 92
            image.height = 46
            sheet.add_image(image, "AK2")
        except Exception:
            pass


def _write_if_empty(sheet: Worksheet, coordinate: str, value: str) -> None:
    cell = sheet[coordinate]
    if cell.value in {None, ""}:
        cell.value = value


def _mapped_columns(mapping: ColumnMapping) -> list[int]:
    return sorted({column for columns in mapping.values() for column in columns})


def _normalize(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).lower().replace("\n", " ").split()).strip(" :")


def _amount(value: Decimal | None) -> int | float | None:
    if value is None:
        return None
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def _unit_price(line: BudgetLine) -> Decimal | None:
    if line.totaal_prijs_per_regel is not None and line.hoeveelheid not in {None, 0}:
        try:
            return line.totaal_prijs_per_regel / line.hoeveelheid
        except (InvalidOperation, ZeroDivisionError):
            return line.eenheidsprijs
    return line.eenheidsprijs
