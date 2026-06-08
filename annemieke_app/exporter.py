from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from decimal import Decimal, InvalidOperation

from .models import BudgetLine, IncomingDocument
from .normalizer import is_noise_line


HEADERS = [
    "Omschrijving / werkzaamheden",
    "Hvh",
    "Ehd",
    "Norm / arbeid",
    "Uren",
    "Materiaal",
    "Materieel",
    "O.A.",
    "Eindprijs",
    "Eenheidsprijs",
]

CONTROL_MODEL_HEADERS = [
    "Hoofdstuk",
    "Post",
    "Omschrijving / werkzaamheden",
    "Hvh",
    "Ehd",
    "Norm",
    "Uren",
    "Materiaal",
    "Materieel",
    "Onderaannemer",
    "Eindprijs",
    "Pagina",
    "Score",
]


def budget_document_to_xlsx(document: IncomingDocument) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Begroting"

    sheet["A1"] = "Project"
    sheet["B1"] = document.project_name or ""
    sheet["A2"] = "Bronbestand"
    sheet["B2"] = document.original_filename
    sheet["A3"] = "Status"
    sheet["B3"] = document.status
    sheet["A4"] = "Originele input"
    sheet["B4"] = document.source_total_amount
    sheet["A5"] = "Regeltotaal"
    sheet["B5"] = _line_total(document.budget_lines)
    for row in range(1, 6):
        sheet[f"A{row}"].font = Font(bold=True, color="1F4E78")
        sheet[f"B{row}"].font = Font(bold=True)
    sheet["B4"].number_format = u'€ #,##0.00'
    sheet["B5"].number_format = u'€ #,##0.00'

    start_row = 7
    sheet.append([])
    sheet.append(HEADERS)
    for cell in sheet[start_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style="thin", color="7F7F7F"))

    for line in document.budget_lines:
        sheet.append(
            [
                line.omschrijving_werkzaamheden,
                line.hoeveelheid,
                line.eenheid,
                line.norm_arbeid,
                line.uren,
                line.materiaal,
                line.materieel,
                line.onderaannemer,
                _line_total_value(line),
                _unit_price(line),
            ]
        )

    widths = [42, 12, 8, 12, 10, 13, 13, 13, 18, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=start_row + 1):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="right")
        for cell in row:
            cell.border = Border(bottom=Side(style="hair", color="D9D9D9"))

    for row in sheet.iter_rows(min_row=start_row + 1, min_col=2, max_col=10):
        for cell in row:
            if cell.column == 3:
                continue
            cell.number_format = '#,##0.00'

    last_row = max(start_row + len(document.budget_lines), start_row)
    table = Table(displayName="BegrotingRegels", ref=f"A{start_row}:J{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = f"A{start_row + 1}"
    sheet.auto_filter.ref = f"A{start_row}:J{last_row}"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def control_model_document_to_xlsx(document: IncomingDocument) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Controlemodel"

    sheet["A1"] = "Project"
    sheet["B1"] = document.project_name or (document.project.name if document.project else "")
    sheet["A2"] = "Bronbestand"
    sheet["B2"] = document.original_filename
    sheet["A3"] = "Model"
    sheet["B3"] = "Begroting beoordeling"
    sheet["A4"] = "Originele input"
    sheet["B4"] = document.source_total_amount
    sheet["A5"] = "Regeltotaal"
    sheet["B5"] = _line_total(document.budget_lines)
    for row in range(1, 6):
        sheet[f"A{row}"].font = Font(bold=True, color="1F4E78")
        sheet[f"B{row}"].font = Font(bold=True)
    sheet["B4"].number_format = u'€ #,##0.00'
    sheet["B5"].number_format = u'€ #,##0.00'

    start_row = 7
    sheet.append([])
    sheet.append(CONTROL_MODEL_HEADERS)
    for cell in sheet[start_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style="thin", color="7F7F7F"))

    for line in document.budget_lines:
        row_number = sheet.max_row + 1
        sheet.append(
            [
                line.hoofdstuk_code or "",
                line.post_code or "",
                line.omschrijving_werkzaamheden,
                _display_amount(line.hoeveelheid),
                line.eenheid or "",
                line.norm_arbeid,
                line.uren,
                line.materiaal,
                line.materieel,
                line.onderaannemer,
                _line_total_value(line),
                line.bron_pagina,
                line.confidence,
            ]
        )
        if line.regel_type in {"hoofdstuk", "post"}:
            for cell in sheet[row_number]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="4BAFC9" if line.regel_type == "hoofdstuk" else "2C5B88")

    widths = [14, 14, 56, 10, 8, 12, 12, 14, 14, 16, 14, 10, 10]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=start_row + 1):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
        for cell in row:
            cell.border = Border(bottom=Side(style="hair", color="D9D9D9"))
        for cell in row[5:7]:
            cell.number_format = '#,##0.00'
        for cell in row[7:11]:
            cell.number_format = u'€ #,##0.00'
        row[12].number_format = '0"%"'

    last_row = max(start_row + len(document.budget_lines), start_row)
    table = Table(displayName="Controlemodel", ref=f"A{start_row}:M{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = f"A{start_row + 1}"
    sheet.auto_filter.ref = f"A{start_row}:M{last_row}"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def selected_budget_lines_to_xlsx(lines: list[BudgetLine]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Selectie"
    headers = [
        "Datum",
        "Project",
        "Projectnr",
        "Relatie",
        "Document",
        "Omschrijving",
        "Hvh",
        "Ehd",
        "Eenheidsprijs",
        "Totaal",
        "Score",
        "Status",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    for line in lines:
        document = line.document
        project = document.project
        sheet.append(
            [
                document.created_at.strftime("%d-%m-%Y") if document.created_at else "",
                project.name if project else document.project_name or "",
                project.project_number if project else "",
                project.client.name if project and project.client else "",
                document.original_filename,
                line.omschrijving_werkzaamheden,
                _display_amount(line.hoeveelheid),
                line.eenheid or "",
                _unit_price(line),
                _line_total_value(line),
                line.confidence,
                document.status,
            ]
        )

    widths = [13, 26, 16, 24, 32, 48, 10, 8, 14, 14, 10, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=2):
        row[5].alignment = Alignment(wrap_text=True, vertical="top")
        for cell in row:
            cell.border = Border(bottom=Side(style="hair", color="D9D9D9"))
        row[8].number_format = u'€ #,##0.00'
        row[9].number_format = u'€ #,##0.00'
        row[10].number_format = '0"%"'

    last_row = max(len(lines) + 1, 1)
    table = Table(displayName="RaadplegenSelectie", ref=f"A1:L{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:L{last_row}"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _unit_price(line: BudgetLine) -> Decimal | None:
    if line.totaal_prijs_per_regel is not None and line.hoeveelheid not in {None, 0}:
        try:
            return line.totaal_prijs_per_regel / line.hoeveelheid
        except (InvalidOperation, ZeroDivisionError):
            return line.eenheidsprijs
    return line.eenheidsprijs or _price_component_total(line)


def _line_total_value(line: BudgetLine) -> Decimal | None:
    unit_price = _unit_price(line)
    if line.hoeveelheid not in {None, 0} and unit_price is not None:
        try:
            return Decimal(str(line.hoeveelheid)) * Decimal(str(unit_price))
        except InvalidOperation:
            return line.totaal_prijs_per_regel
    return line.totaal_prijs_per_regel or unit_price


def _line_total(lines: list[BudgetLine]) -> Decimal:
    usable_lines = [
        line
        for line in lines
        if _line_total_value(line) is not None
        and line.regel_type not in {"hoofdstuk", "post", "subtotaal", "totaal"}
        and not is_noise_line(line.omschrijving_werkzaamheden)
    ]
    if not usable_lines:
        usable_lines = [line for line in lines if _line_total_value(line) is not None]

    total = Decimal("0")
    for line in usable_lines:
        total += Decimal(str(_line_total_value(line) or 0))
    return total


def _price_component_total(line: BudgetLine) -> Decimal | None:
    total = Decimal("0")
    has_component = False
    for value in (line.materiaal, line.materieel, line.onderaannemer):
        if value is None:
            continue
        total += Decimal(str(value))
        has_component = True
    return total if has_component else None


def _display_amount(value: Decimal | None) -> int | float | None:
    if value is None:
        return None
    if value == value.to_integral_value():
        return int(value)
    return float(value.quantize(Decimal("0.01")))
