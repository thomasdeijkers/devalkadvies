from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.engine import make_url
from sqlalchemy.exc import ArgumentError

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

_database_url = os.getenv("DATABASE_URL", "")
if _database_url:
    try:
        make_url(_database_url)
    except ArgumentError as exc:
        raise SystemExit(
            "DATABASE_URL is geen geldige database-url. Zet hem eerst goed in PowerShell, bijvoorbeeld:\n"
            "$env:DATABASE_URL='postgresql+psycopg://gebruiker:wachtwoord@host:5432/devalkadvies'\n"
            "Let op: speciale tekens in het wachtwoord moeten URL-encoded zijn, bijvoorbeeld @ = %40."
        ) from exc
    if "<" in _database_url or ">" in _database_url:
        raise SystemExit("DATABASE_URL bevat nog placeholder-tekst met <...>. Vul de echte database-url in.")

from annemieke_app.config import settings  # noqa: E402
from annemieke_app.database import SessionLocal, create_db  # noqa: E402
from annemieke_app.kengetallen import import_reference_lines  # noqa: E402
from annemieke_app.models import PriceIndexSeries, PriceIndexValue, ReferenceDataset, ReferenceLine  # noqa: E402
from annemieke_app.normalizer import apply_normalization, is_noise_line  # noqa: E402


DEFAULT_SOURCE_DIR = ROOT / "Hulp bestanden" / "Kengetallen"
PRICE_INDEX_SERIES_NAME = "Nieuwbouwwoningen outputprijsindex bouwkosten"
PRICE_INDEX_SOURCE = "CBS Prijsindex bouwkosten excl. BTW | Index"
SUMMARY_LABELS = {"laag", "gemiddeld", "hoog"}
META_HEADERS = {"fase", "peildatum", "periode", "bdb indexering"}
REFERENCE_INDEX_TYPE = "kengetal_index"
REFERENCE_PROJECT_SHEET_TYPE = "kengetal_bronregel"
EXCLUDED_PROJECT_SHEETS = {"indexen", "vormfactoren", "kengetallen geindexeerd", "kengetallen geïndexeerd"}
CATEGORY_HEADERS = {
    "algemeen": "Algemeen",
    "sloopwerk": "Sloopwerk",
    "fundering": "Fundering",
    "skelet": "Skelet",
    "daken": "Daken",
    "gevel": "Gevel",
    "binnenwanden": "Binnenwanden",
    "vloerafwerking": "Vloerafwerking",
    "trappen": "Trappen",
    "plafondafwerking": "Plafondafwerking",
    "vaste inrichting": "Vaste inrichting",
    "terrein": "Terrein",
    "w installatie": "W installatie",
    "e installatie": "E installatie",
    "t installatie": "T installatie",
    "bouwkundig": "Bouwkundig",
    "installatie": "Installatie",
}


def main() -> None:
    parser = argparse.ArgumentParser(description="Laad DeValk kengetallenbestanden eenmalig in de database.")
    parser.add_argument(
        "paths",
        nargs="*",
        type=Path,
        help="Excelbestanden of mappen. Leeg = Hulp bestanden/Kengetallen.",
    )
    args = parser.parse_args()
    files = _collect_files(args.paths or [DEFAULT_SOURCE_DIR])
    if not files:
        raise SystemExit(f"Geen Excelbestanden gevonden in {DEFAULT_SOURCE_DIR}")

    create_db()
    with SessionLocal() as session:
        total_lines = 0
        for path in files:
            line_count = import_file(session, path)
            total_lines += line_count
            print(f"{path.name}: {line_count} kengetalregels geladen")
        print(f"Klaar: {len(files)} bestand(en), {total_lines} regels")


def import_file(session, path: Path) -> int:
    _delete_existing_dataset(session, path.name)
    dataset = ReferenceDataset(
        name=path.stem,
        original_filename=path.name,
        stored_filename=_store_source_file(path),
        source="kengetallen importscript",
        status="active",
        notes="Eenmalig ingeladen bronbestand voor historische kengetallen en normalisatie.",
    )
    session.add(dataset)
    session.flush()

    lines = _kengetallen_index_lines(path, dataset)
    if lines:
        lines.extend(_project_sheet_source_lines(path, dataset, start_line_number=len(lines) + 1))
    else:
        lines = _generic_reference_lines(path, dataset)
    for line in lines:
        session.add(line)
    if lines:
        apply_normalization(session, lines)
    index_count = _import_price_index_values(session, path)
    if index_count:
        dataset.notes = (
            "Eenmalig ingeladen bronbestand voor historische kengetallen en normalisatie. "
            f"Indexblad verwerkt: {index_count} periodes."
        )
    session.commit()
    return len(lines)


def _collect_files(paths: list[Path]) -> list[Path]:
    files: list[Path] = []
    for path in paths:
        resolved = path if path.is_absolute() else ROOT / path
        if resolved.is_dir():
            files.extend(sorted(resolved.glob("*.xlsx")))
            files.extend(sorted(resolved.glob("*.xlsm")))
        elif resolved.suffix.lower() in {".xlsx", ".xlsm"} and resolved.exists():
            files.append(resolved)
    return sorted({file.resolve() for file in files})


def _delete_existing_dataset(session, original_filename: str) -> None:
    existing = session.scalar(
        select(ReferenceDataset).where(ReferenceDataset.original_filename == original_filename).limit(1)
    )
    if existing is not None:
        session.delete(existing)
        session.flush()


def _store_source_file(path: Path) -> str:
    settings.upload_dir.mkdir(parents=True, exist_ok=True)
    stored_filename = f"kengetallen-{uuid4().hex}{path.suffix.lower()}"
    target = settings.upload_dir / stored_filename
    target.write_bytes(path.read_bytes())
    return stored_filename


def _kengetallen_index_lines(path: Path, dataset: ReferenceDataset) -> list[ReferenceLine]:
    workbook = load_workbook(path, data_only=True, keep_vba=path.suffix.lower() == ".xlsm")
    sheet = _find_index_sheet(workbook.worksheets)
    if sheet is None:
        return []
    blocks = _find_header_blocks(sheet)
    sheet_cache: dict[tuple[str, str], str | None] = {}
    lines: list[ReferenceLine] = []
    for block_index, block in enumerate(blocks):
        next_row = blocks[block_index + 1]["row"] if block_index + 1 < len(blocks) else sheet.max_row + 1
        project_column = block["project_column"]
        variant_column = project_column + 1
        for row in range(block["row"] + 1, next_row):
            project_name = _text(sheet.cell(row=row, column=project_column).value)
            if not project_name or project_name.lower() in SUMMARY_LABELS:
                continue
            variant = _text(sheet.cell(row=row, column=variant_column).value) or None
            phase = _text(_cell(sheet, row, block["columns"].get("fase")))
            period = _text(_cell(sheet, row, block["columns"].get("periode")))
            price_date = _date(_cell(sheet, row, block["columns"].get("peildatum")))
            index_factor = _decimal(_cell(sheet, row, block["columns"].get("bdb indexering")))
            lookup_key = (project_name, variant or "")
            if lookup_key not in sheet_cache:
                sheet_cache[lookup_key] = _match_project_sheet(workbook.worksheets, project_name, variant)
            project_sheet_name = sheet_cache[lookup_key]
            for category_key, category_label in CATEGORY_HEADERS.items():
                column = block["columns"].get(category_key)
                value = _decimal(_cell(sheet, row, column))
                if value is None:
                    continue
                description = f"{category_label} - {block['section']}"
                lines.append(
                    ReferenceLine(
                        dataset_id=dataset.id,
                        line_number=len(lines) + 1,
                        regel_type=REFERENCE_INDEX_TYPE,
                        niveau=0,
                        hoofdstuk_code=category_key,
                        hoofdstuk_omschrijving=block["section"],
                        post_code=category_key,
                        project_name=_limit_text(project_name, 180),
                        relation_name=_limit_text(variant, 180),
                        document_date=price_date,
                        phase=_limit_text(phase, 120),
                        period=_limit_text(period, 40),
                        bdb_indexering=index_factor,
                        project_sheet_name=_limit_text(project_sheet_name, 180),
                        source_row=row,
                        omschrijving_werkzaamheden=description,
                        hoeveelheid=Decimal("1"),
                        eenheid="m2 bvo",
                        norm_arbeid=index_factor,
                        totaal_prijs_per_regel=value,
                        eenheidsprijs=value,
                        confidence=100,
                        raw_text=(
                            f"{path.name}|{sheet.title}|rij:{row}|categorie:{category_key}|"
                            f"fase:{phase}|periode:{period}"
                        ),
                    )
                )
    return lines


def _import_price_index_values(session, path: Path) -> int:
    workbook = load_workbook(path, data_only=True, keep_vba=path.suffix.lower() == ".xlsm")
    sheet = next((ws for ws in workbook.worksheets if _key(ws.title) == "indexen"), None)
    if sheet is None:
        return 0

    period_column = None
    index_column = None
    header_row = None
    for row in range(1, min(sheet.max_row, 30) + 1):
        for column in range(1, min(sheet.max_column, 12) + 1):
            label = _key(sheet.cell(row=row, column=column).value)
            if label == "periode":
                period_column = column
            elif label == "index":
                index_column = column
        if period_column and index_column:
            header_row = row
            break
    if not header_row:
        return 0

    series = session.scalar(select(PriceIndexSeries).where(PriceIndexSeries.name == PRICE_INDEX_SERIES_NAME))
    if series is None:
        series = PriceIndexSeries(name=PRICE_INDEX_SERIES_NAME)
        session.add(series)
        session.flush()
    source_text = _index_sheet_source_text(sheet, header_row)
    series.description = source_text or "CBS Prijsindex bouwkosten excl. BTW, ingeladen uit het DeValk kengetallenbestand."
    series.source = f"{path.name} > {sheet.title}"
    series.provider = "excel"
    series.period_field = "Periode"
    series.value_field = "Index"
    series.last_synced_at = datetime.now()
    series.values.clear()
    session.flush()

    count = 0
    for row in range(header_row + 1, sheet.max_row + 1):
        period = _text(sheet.cell(row=row, column=period_column).value)
        index_value = _decimal(sheet.cell(row=row, column=index_column).value)
        effective_date = _period_start(period)
        if not period or index_value is None or effective_date is None:
            continue
        session.add(
            PriceIndexValue(
                series_id=series.id,
                effective_date=effective_date,
                index_value=index_value,
                notes=period,
                source_reference=f"{path.name} > {sheet.title} rij {row} | {PRICE_INDEX_SOURCE}",
            )
        )
        count += 1
    return count


def _index_sheet_source_text(sheet: Worksheet, header_row: int) -> str:
    values: list[str] = []
    for row in range(1, max(header_row, 1)):
        for column in range(1, min(sheet.max_column, 10) + 1):
            text = _text(sheet.cell(row=row, column=column).value)
            if text and _key(text) not in {"periode", "index"}:
                values.append(text)
    return " | ".join(dict.fromkeys(values))


def _project_sheet_source_lines(
    path: Path,
    dataset: ReferenceDataset,
    start_line_number: int = 1,
) -> list[ReferenceLine]:
    workbook = load_workbook(path, data_only=True, keep_vba=path.suffix.lower() == ".xlsm")
    sheet_metadata = _index_metadata_by_sheet(workbook)
    lines: list[ReferenceLine] = []
    try:
        for sheet in workbook.worksheets:
            if sheet.sheet_state != "visible" or _key(sheet.title) in EXCLUDED_PROJECT_SHEETS:
                continue
            header = _project_sheet_header(sheet)
            if not header:
                continue
            metadata = sheet_metadata.get(sheet.title, {})
            current_section = ""
            for row in range(header["row"] + 1, sheet.max_row + 1):
                description = _text(_cell(sheet, row, header.get("description")))
                code = _text(_cell(sheet, row, header.get("code")))
                if not description:
                    continue
                quantity = _decimal(_cell(sheet, row, header.get("quantity")))
                unit = _text(_cell(sheet, row, header.get("unit"))) or None
                unit_price = _decimal(_cell(sheet, row, header.get("unit_price")))
                total_price = _decimal(_cell(sheet, row, header.get("total_price")))
                m2bvo_price = _decimal(_cell(sheet, row, header.get("m2bvo_price")))
                if total_price is None and unit_price is not None and quantity not in (None, Decimal("0")):
                    total_price = unit_price * quantity
                if unit_price is None and total_price is not None and quantity not in (None, Decimal("0")):
                    try:
                        unit_price = total_price / quantity
                    except (InvalidOperation, ZeroDivisionError):
                        unit_price = None
                if not any([quantity, unit, unit_price, total_price, m2bvo_price]):
                    clean_description = description.strip(" -")
                    if clean_description and len(clean_description) <= 90:
                        current_section = clean_description
                source_text = f"{path.name}|{sheet.title}|rij:{row}|bronregel"
                if m2bvo_price is not None:
                    source_text = f"{source_text}|m2bvo:{m2bvo_price}"
                lines.append(
                    ReferenceLine(
                        dataset_id=dataset.id,
                        line_number=start_line_number + len(lines),
                        regel_type=REFERENCE_PROJECT_SHEET_TYPE,
                        niveau=0,
                        hoofdstuk_code=_limit_text(code, 80),
                        hoofdstuk_omschrijving=_limit_text(current_section, 255),
                        post_code=_limit_text(code, 80),
                        project_name=_limit_text(metadata.get("project_name") or _project_title_from_sheet(sheet), 180),
                        relation_name=_limit_text(metadata.get("variant"), 180),
                        document_date=metadata.get("document_date"),
                        phase=_limit_text(metadata.get("phase"), 120),
                        period=_limit_text(metadata.get("period"), 40),
                        bdb_indexering=metadata.get("bdb_indexering"),
                        project_sheet_name=_limit_text(sheet.title, 180),
                        source_row=row,
                        omschrijving_werkzaamheden=description,
                        hoeveelheid=quantity,
                        eenheid=_limit_text(unit, 40),
                        totaal_prijs_per_regel=total_price,
                        eenheidsprijs=unit_price,
                        bron_pagina=None,
                        confidence=100 if unit_price is not None or total_price is not None else 70,
                        raw_text=source_text,
                    )
                )
    finally:
        workbook.close()
    return lines


def _index_metadata_by_sheet(workbook) -> dict[str, dict[str, Any]]:
    sheet = _find_index_sheet(workbook.worksheets)
    if sheet is None:
        return {}
    metadata: dict[str, dict[str, Any]] = {}
    blocks = _find_header_blocks(sheet)
    for block_index, block in enumerate(blocks):
        next_row = blocks[block_index + 1]["row"] if block_index + 1 < len(blocks) else sheet.max_row + 1
        project_column = block["project_column"]
        variant_column = project_column + 1
        for row in range(block["row"] + 1, next_row):
            project_name = _text(sheet.cell(row=row, column=project_column).value)
            if not project_name or project_name.lower() in SUMMARY_LABELS:
                continue
            variant = _text(sheet.cell(row=row, column=variant_column).value) or None
            project_sheet_name = _match_project_sheet(workbook.worksheets, project_name, variant)
            if not project_sheet_name or project_sheet_name in metadata:
                continue
            metadata[project_sheet_name] = {
                "project_name": project_name,
                "variant": variant,
                "phase": _text(_cell(sheet, row, block["columns"].get("fase"))) or None,
                "period": _text(_cell(sheet, row, block["columns"].get("periode"))) or None,
                "document_date": _date(_cell(sheet, row, block["columns"].get("peildatum"))),
                "bdb_indexering": _decimal(_cell(sheet, row, block["columns"].get("bdb indexering"))),
            }
    return metadata


def _project_sheet_header(sheet: Worksheet) -> dict[str, int] | None:
    for row in range(1, min(sheet.max_row, 120) + 1):
        columns: dict[str, int] = {}
        for column in range(1, min(sheet.max_column, 14) + 1):
            token = _header_token(sheet.cell(row=row, column=column).value)
            if token in {"code", "codering"}:
                columns["code"] = column
            elif token in {"onderdeel", "omschrijving"}:
                columns["description"] = column
            elif token in {"hheid", "hvheid", "hvh"}:
                columns["quantity"] = column
            elif token in {"eheid", "ehd", "eenheid"}:
                columns["unit"] = column
            elif token in {"euroeenheid", "eureenheid", "eenheidsprijs"}:
                columns["unit_price"] = column
            elif token in {"eurototaal", "eurtotaal", "totaal"}:
                columns["total_price"] = column
            elif token in {"eurom2bvo", "eurm2bvo", "m2bvo"}:
                columns["m2bvo_price"] = column
        if "description" in columns and ("unit_price" in columns or "total_price" in columns or "m2bvo_price" in columns):
            columns["row"] = row
            columns.setdefault("code", max(1, columns["description"] - 1))
            return columns
    return None


def _project_title_from_sheet(sheet: Worksheet) -> str:
    for row in range(1, min(sheet.max_row, 12) + 1):
        values = [_text(sheet.cell(row=row, column=column).value) for column in range(1, min(sheet.max_column, 10) + 1)]
        line = " ".join(value for value in values if value)
        match = re.search(r"Project\s*&\s*Locatie\s*:?\s*(.+)", line, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return sheet.title


def _header_token(value: Any) -> str:
    text = _text(value).lower()
    text = text.replace("€", "euro").replace("/", "").replace("-", "")
    return re.sub(r"[^a-z0-9]+", "", text)


def _limit_text(value: Any, max_length: int) -> str | None:
    text = _text(value)
    if not text:
        return None
    return text[:max_length]


def _generic_reference_lines(path: Path, dataset: ReferenceDataset) -> list[ReferenceLine]:
    imported_lines = import_reference_lines(path)
    lines: list[ReferenceLine] = []
    for imported in imported_lines:
        if is_noise_line(imported.omschrijving_werkzaamheden):
            continue
        lines.append(
            ReferenceLine(
                dataset_id=dataset.id,
                line_number=len(lines) + 1,
                regel_type=imported.regel_type,
                niveau=imported.niveau,
                hoofdstuk_code=imported.hoofdstuk_code,
                hoofdstuk_omschrijving=imported.hoofdstuk_omschrijving,
                post_code=imported.post_code,
                project_name=imported.project_name,
                relation_name=imported.relation_name,
                document_date=imported.document_date,
                omschrijving_werkzaamheden=imported.omschrijving_werkzaamheden,
                hoeveelheid=imported.hoeveelheid,
                eenheid=imported.eenheid,
                norm_arbeid=imported.norm_arbeid,
                uren=imported.uren,
                materiaal=imported.materiaal,
                materieel=imported.materieel,
                onderaannemer=imported.onderaannemer,
                totaal_prijs_per_regel=imported.totaal_prijs_per_regel,
                eenheidsprijs=imported.eenheidsprijs,
                bron_pagina=imported.bron_pagina,
                confidence=imported.confidence,
                raw_text=imported.raw_text,
            )
        )
    return lines


def _find_index_sheet(sheets: list[Worksheet]) -> Worksheet | None:
    scored: list[tuple[int, Worksheet]] = []
    for sheet in sheets:
        score = 10 if "kengetallen" in sheet.title.lower() else 0
        for row in range(1, min(sheet.max_row, 80) + 1):
            labels = {_key(sheet.cell(row=row, column=column).value) for column in range(1, min(sheet.max_column, 45) + 1)}
            score += len(labels & META_HEADERS) * 3
            score += len(labels & set(CATEGORY_HEADERS)) * 2
        scored.append((score, sheet))
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored[0][1] if scored and scored[0][0] >= 20 else None


def _match_project_sheet(sheets: list[Worksheet], project_name: str, variant: str | None) -> str | None:
    candidates = [
        sheet
        for sheet in sheets
        if _key(sheet.title) not in EXCLUDED_PROJECT_SHEETS
    ]
    project_key = _compact_key(project_name)
    variant_key = _compact_key(variant or "")
    best_score = 0
    best_title = None
    for sheet in candidates:
        title_key = _compact_key(sheet.title)
        score = 0
        if title_key and title_key in project_key:
            score += 50
        if project_key and project_key in title_key:
            score += 50
        if variant_key and variant_key in title_key:
            score += 15
        if score == 0:
            for row in range(1, min(sheet.max_row, 12) + 1):
                row_text = " ".join(_text(sheet.cell(row=row, column=column).value) for column in range(1, min(sheet.max_column, 12) + 1))
                row_key = _compact_key(row_text)
                if project_key and project_key in row_key:
                    score += 40
                    break
                if variant_key and variant_key in row_key:
                    score += 15
        if score > best_score:
            best_score = score
            best_title = sheet.title
    return best_title if best_score >= 20 else None


def _find_header_blocks(sheet: Worksheet) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    current_section = ""
    for row in range(1, min(sheet.max_row, 1000) + 1):
        columns: dict[str, int] = {}
        for column in range(1, sheet.max_column + 1):
            label = _key(sheet.cell(row=row, column=column).value)
            if label in META_HEADERS or label in CATEGORY_HEADERS:
                columns[label] = column
        category_count = len(set(columns) & set(CATEGORY_HEADERS))
        if "fase" in columns and "periode" in columns and category_count >= 4:
            project_column = max(1, columns["fase"] - 2)
            section = _text(sheet.cell(row=row, column=project_column).value) or current_section or f"Blok rij {row}"
            current_section = section
            blocks.append(
                {
                    "row": row,
                    "section": section,
                    "columns": columns,
                    "project_column": project_column,
                }
            )
        else:
            first_values = [_text(sheet.cell(row=row, column=column).value) for column in range(1, min(sheet.max_column, 5) + 1)]
            first_text = next((value for value in first_values if value), "")
            if first_text and first_text.lower() not in SUMMARY_LABELS:
                current_section = first_text
    return blocks


def _key(value: Any) -> str:
    return re.sub(r"\s+", " ", _text(value).lower().replace("bdb-", "bdb")).strip(" :")


def _compact_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _text(value).lower())


def _cell(sheet: Worksheet, row: int, column: int | None) -> Any:
    if not column:
        return None
    return sheet.cell(row=row, column=column).value


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    text = str(value).strip().lower()
    if text in {"x", "-", "n.v.t.", "nvt"}:
        return None
    cleaned = text.replace("€", "").replace(" ", "")
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = _text(value)
    for pattern in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def _period_start(period: str) -> datetime | None:
    match = re.search(r"(\d{4})\s*[-_ ]?\s*q([1-4])", period.lower())
    if match:
        year = int(match.group(1))
        quarter = int(match.group(2))
        return datetime(year, 1 + (quarter - 1) * 3, 1)
    return _date(period)


if __name__ == "__main__":
    main()
